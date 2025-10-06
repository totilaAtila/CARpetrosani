# vizualizare_trimestriala.py

"""
Modul pentru widget‑ul de vizualizare trimestrială.
Selectare trimestru și an prin ComboBox, cu aceleași stiluri ca în modul anual.
"""

import os
import sqlite3
import sys
from datetime import datetime

from PyQt5 import QtWidgets
from PyQt5.QtWidgets import (
    QApplication, QWidget, QLabel, QPushButton, QVBoxLayout, QHBoxLayout,
    QScrollArea, QMessageBox, QTableWidget, QTableWidgetItem, QComboBox,
    QHeaderView, QFileDialog
)
from PyQt5.QtGui import QCursor, QColor, QBrush
from PyQt5.QtCore import Qt, QEvent
from reportlab.lib.pagesizes import A4, landscape
from reportlab.pdfgen import canvas
from reportlab.lib.units import mm, cm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from PyQt5.QtGui import QBrush, QColor
from reportlab.lib.colors import HexColor, red, black

# Import pentru export Excel
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from utils import ProgressDialog

if getattr(sys, 'frozen', False):
    # Daca aplicatia este impachetata (ruleaza din executabil)
    # os.path.dirname(sys.executable) va fi directorul care contine executabilul (in onedir mode)
    # Sau directorul temporar in onefile mode (dar onefile e mai complicat pt fisiere externe)
    # Presupunem onedir mode si bazele de date sunt langa exe
    BASE_RESOURCE_PATH = os.path.dirname(sys.executable)
else:
    # Daca ruleaza din scriptul sursa Python (in timpul dezvoltarii)
    # __file__ este calea catre scriptul curent (ex. ui/dividende.py)
    # os.path.dirname(__file__) este directorul scriptului curent (ex. ui)
    # os.path.join(..., "..") urca un nivel (la directorul radacina al proiectului)
    BASE_RESOURCE_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..")

# Definim caile catre bazele de date relative la directorul de resurse
DB_MEMBRII = os.path.join(BASE_RESOURCE_PATH, "MEMBRII.db")
DB_DEPCRED = os.path.join(BASE_RESOURCE_PATH, "DEPCRED.db")
DB_ACTIVI = os.path.join(BASE_RESOURCE_PATH, "ACTIVI.db")
DB_INACTIVI = os.path.join(BASE_RESOURCE_PATH, "INACTIVI.db")
DB_CHITANTE = os.path.join(BASE_RESOURCE_PATH, "CHITANTE.db")
DB_LICHIDATI = os.path.join(BASE_RESOURCE_PATH, "LICHIDATI.db")

# ----------------------------------------------------------------------------
#  Înregistrare fonturi PDF  (DejaVu > Arial > Helvetica)
# ----------------------------------------------------------------------------
try:
    # MODIFICARE - Folosim BASE_RESOURCE_PATH în loc de BASE_DIR pentru consistență
    ARIAL = os.path.join(BASE_RESOURCE_PATH, "Arial.ttf")

    # Definim posibile căi pentru fonturile DejaVu
    DEJAVU_FILES = [
        "DejaVuSans.ttf",
        "DejaVuSans-Bold.ttf",
        "DejaVuBoldSans.ttf"
    ]

    # Căutăm fonturile DejaVu în directorul de resurse
    DEJAVU = next((os.path.join(BASE_RESOURCE_PATH, f) for f in DEJAVU_FILES
                   if os.path.exists(os.path.join(BASE_RESOURCE_PATH, f))), None)

    # Înregistrăm fonturile disponibile
    if os.path.exists(ARIAL):
        pdfmetrics.registerFont(TTFont("Arial", ARIAL))
    if DEJAVU and os.path.exists(DEJAVU):
        pdfmetrics.registerFont(TTFont("DejaVu", DEJAVU))

    # Setăm fontul implicit
    DEFAULT_FONT = (
        "DejaVu" if "DejaVu" in pdfmetrics.getRegisteredFontNames() else
        "Arial" if "Arial" in pdfmetrics.getRegisteredFontNames() else
        "Helvetica"
    )
except Exception:
    DEFAULT_FONT = "Helvetica"


class VizualizareTrimestrialaWidget(QWidget):
    """Widget pentru vizualizarea situației financiare pe un trimestru."""

    def __init__(self):
        super().__init__()
        self.widgets_cu_cursor_mana: list[QtWidgets.QWidget] = []
        self.date_trimestru: list[dict[str, float]] = []
        # MODIFICARE - Adăugare atribute pentru sortare
        self.sort_order = Qt.AscendingOrder  # Ordinea sortării: implicit ascendentă
        self.sort_column = 2  # Coloana implicită de sortare (nume prenume)
        self._init_ui()

    def _init_ui(self):
        principal = QVBoxLayout(self)

        # Selector trimestru/an -----------------------------------------------
        selector_layout = QHBoxLayout()
        selector_layout.addStretch(1)
        selector_layout.setSpacing(10)

        inner = QHBoxLayout()
        lbl_trim = QLabel("Selectare trimestru:")
        inner.addWidget(lbl_trim)
        inner.setSpacing(5)

        self.combo_trimestru = QComboBox()
        self.trimestre = {
            "Trimestrul 1 (Ian–Mar)": [1, 2, 3],
            "Trimestrul 2 (Apr–Iun)": [4, 5, 6],
            "Trimestrul 3 (Iul–Sep)": [7, 8, 9],
            "Trimestrul 4 (Oct–Dec)": [10, 11, 12]
        }
        self.combo_trimestru.addItems(self.trimestre.keys())
        self.combo_trimestru.setCurrentIndex((datetime.now().month - 1) // 3)
        self.combo_trimestru.setMinimumWidth(160)
        inner.addWidget(self.combo_trimestru)

        lbl_an = QLabel("Selectare an:")
        inner.addWidget(lbl_an)
        inner.setSpacing(5)

        self.combo_an = QComboBox()
        cur_year = datetime.now().year
        ani = [str(y) for y in range(cur_year - 25, cur_year + 5)]
        self.combo_an.addItems(ani)
        self.combo_an.setCurrentText(str(cur_year))
        self.combo_an.setMinimumWidth(80)
        inner.addWidget(self.combo_an)

        self.btn_show = QPushButton("Afișează trimestru")
        self.btn_show.clicked.connect(self.afiseaza_trimestru)
        inner.addWidget(self.btn_show)

        self.btn_total = QPushButton("Afișare total trimestru")
        self.btn_total.setEnabled(False)
        self.btn_total.clicked.connect(self.afiseaza_totaluri)
        inner.addWidget(self.btn_total)

        self.btn_export = QPushButton("Exportă PDF")
        self.btn_export.setEnabled(False)
        self.btn_export.clicked.connect(self.exporta_pdf)
        inner.addWidget(self.btn_export)

        # MODIFICARE - Adăugare buton export Excel
        self.btn_export_excel = QPushButton("Exportă Excel")
        self.btn_export_excel.setEnabled(False)
        self.btn_export_excel.clicked.connect(self.exporta_excel)
        inner.addWidget(self.btn_export_excel)

        selector_layout.addLayout(inner)
        selector_layout.addStretch(1)
        principal.addLayout(selector_layout)

        # Tabel ---------------------------------------------------------------
        self.tabel = QTableWidget()
        self.tabel.setColumnCount(10)
        self.tabel.setHorizontalHeaderLabels([
            "Luna-An", "Număr\nFișă", "Nume\nPrenume", "Dobândă", "Rată\nÎmprumut",
            "Sold\nÎmprumut", "Cotizație", "Retragere\nFS", "Sold\nDepunere", "Total\nde Plată"
        ])
        column_widths = [70, 60, 200, 80, 100, 100, 100, 100, 100, 120]
        for i, w in enumerate(column_widths):
            self.tabel.setColumnWidth(i, w)

        # MODIFICARE - Configurare sortare
        self.tabel.setSortingEnabled(False)  # Dezactivăm sortarea automată pentru a o gestiona manual
        header = self.tabel.horizontalHeader()
        header.sectionClicked.connect(self.on_header_clicked)  # Conectăm semnalul de click pentru sortare

        self.tabel.setAlternatingRowColors(True)

        scr = QScrollArea()
        scr.setWidgetResizable(True)
        scr.setWidget(self.tabel)
        principal.addWidget(scr)

        self._aplica_stiluri()

        # Cursor de mână ------------------------------------------------------
        self.widgets_cu_cursor_mana = [
            self.combo_trimestru, self.combo_an,
            self.btn_show, self.btn_total, self.btn_export,
            self.btn_export_excel  # MODIFICARE - Adăugat butonul Excel
        ]
        for w in self.widgets_cu_cursor_mana:
            w.installEventFilter(self)

    def _aplica_stiluri(self) -> None:
        self.setStyleSheet(
            """
            QWidget {font-family:Arial; font-size:10pt; background:#f8f8f8;}
            QPushButton {
                background:#90EE90; color:#000; border:1px solid #60c060;
                border-radius:5px; padding:8px 16px; font-weight:bold;
                margin: 5px;
            }
            QPushButton:hover {background:#77dd77;}\
            QPushButton:pressed {background:#60c060;}\
            QPushButton:disabled {background:#d0e0d0; color:#808080; border:1px solid #b0c0b0;}\
            QHeaderView::section {background-color:#dce8ff; color:#333; padding:6px; border:1px solid #c0c8d0; font-weight:bold;}\
            QComboBox {background-color:#f8f8f8; border:1px solid #b0c0b0; border-radius:5px; padding:8px; font-size:10pt; margin:5px;}\
            QComboBox:focus {border-color:#3498db; box-shadow:0 0 0 2px rgba(52,152,219,0.25); outline:none;}\
            QLabel {color:#555; margin-right:5px;}\
            QTableWidget {background:#ffffff; alternate-background-color:#f0f0f0; margin-top:10px;}\
            QTableWidget::item {padding:8px;}"""
        )

    def eventFilter(self, obj, event):  # noqa: N802
        if event.type() == QEvent.Enter and obj in self.widgets_cu_cursor_mana:
            QApplication.setOverrideCursor(Qt.PointingHandCursor)
            return True
        if event.type() == QEvent.Leave and obj in self.widgets_cu_cursor_mana:
            QApplication.restoreOverrideCursor()
            return True
        return super().eventFilter(obj, event)

    # MODIFICARE - Adăugare metode pentru sortare
    def on_header_clicked(self, logical_index):
        """
        Gestionează clicurile pe antetul tabelului pentru a sorta datele.
        """
        if not self.date_trimestru:
            return  # Nu avem date de sortat

        # Inversează ordinea de sortare dacă se face click pe aceeași coloană
        if self.sort_column == logical_index:
            self.sort_order = Qt.DescendingOrder if self.sort_order == Qt.AscendingOrder else Qt.AscendingOrder
        else:
            self.sort_column = logical_index
            self.sort_order = Qt.AscendingOrder

        # Reafișează tabelul cu date sortate
        self.sorteaza_si_afiseaza_date()

    def sorteaza_si_afiseaza_date(self):
        """
        Sortează datele în funcție de coloana și ordinea selectată și actualizează tabelul.
        """
        # Mapare coloane la cheile din dicționarul de date
        column_to_key = {
            0: 'luna',
            1: 'nr_fisa',
            2: 'nume',
            3: 'dobanda',
            4: 'impr_cred',
            5: 'impr_sold',
            6: 'dep_deb',
            7: 'dep_cred',
            8: 'dep_sold',
            9: 'total_plata'
        }

        key = column_to_key.get(self.sort_column)
        if key is None:
            return  # Nu putem sorta după această coloană

        # Sortează datele
        reverse = (self.sort_order == Qt.DescendingOrder)

        # Pentru sortare numerică vs. text
        if key == 'nume':
            self.date_trimestru.sort(key=lambda x: x[key].lower(), reverse=reverse)
        elif key == 'nr_fisa':
            # Asigură-te că nr_fisa este sortat numeric
            self.date_trimestru.sort(key=lambda x: int(x[key]) if str(x[key]).isdigit() else 0, reverse=reverse)
        elif key == 'luna':
            # Sortare specifică pentru luna
            self.date_trimestru.sort(key=lambda x: x[key], reverse=reverse)
        else:
            # Sortare numerică pentru celelalte coloane
            self.date_trimestru.sort(key=lambda x: float(x.get(key, 0)), reverse=reverse)

        # Actualizează tabelul cu datele sortate
        self.actualizeaza_tabel()

    def actualizeaza_tabel(self):
        """
        Actualizează afișarea tabelului cu datele sortate fără a reinteroga baza de date.
        """
        if not self.date_trimestru:
            return

        self.tabel.setRowCount(0)
        self.tabel.setRowCount(len(self.date_trimestru))

        anul = int(self.combo_an.currentText())

        # 6. Populare tabel
        for i, data in enumerate(self.date_trimestru):
            self.tabel.setItem(i, 0, QTableWidgetItem(f"{data['luna']:02d}-{anul}"))
            self.tabel.setItem(i, 1, QTableWidgetItem(str(data['nr_fisa'])))
            self.tabel.setItem(i, 2, QTableWidgetItem(data['nume']))
            self.tabel.setItem(i, 3, QTableWidgetItem(f"{data['dobanda']:.2f}"))

            # Rata împrumut
            item_impr = QTableWidgetItem(
                "NEACHITAT" if data['impr_sold'] > 0 and data['impr_cred'] == 0
                else f"{data['impr_cred']:.2f}"
            )
            if item_impr.text() == "NEACHITAT":
                item_impr.setForeground(QBrush(QColor("red")))
            self.tabel.setItem(i, 4, item_impr)

            # Sold împrumut
            self.tabel.setItem(i, 5, QTableWidgetItem(f"{data['impr_sold']:.2f}"))

            # Depunere lunară
            item_dep = QTableWidgetItem(
                "NEACHITAT" if data['dep_sold'] > 0 and data['dep_deb'] == 0
                else f"{data['dep_deb']:.2f}"
            )
            if item_dep.text() == "NEACHITAT":
                item_dep.setForeground(QBrush(QColor("red")))
            self.tabel.setItem(i, 6, item_dep)

            # Retragere FS și sold depunere
            self.tabel.setItem(i, 7, QTableWidgetItem(f"{data['dep_cred']:.2f}"))
            self.tabel.setItem(i, 8, QTableWidgetItem(f"{data['dep_sold']:.2f}"))

            # Total de plată
            self.tabel.setItem(i, 9, QTableWidgetItem(f"{data['total_plata']:.2f}"))

        # 7. Colorare pe grupuri (toate rândurile unui membru la un loc)
        total_rows = self.tabel.rowCount()
        group = 0
        prev_nr = None
        for r in range(total_rows):
            nr_curent = self.tabel.item(r, 1).text()
            if nr_curent != prev_nr:
                if prev_nr is not None:
                    group += 1
                prev_nr = nr_curent
            bg_color = "#e8f4ff" if (group % 2 == 0) else "#fff5e6"
            brush = QBrush(QColor(bg_color))
            for c in range(self.tabel.columnCount()):
                it = self.tabel.item(r, c)
                if it:
                    it.setBackground(brush)

    def afiseaza_trimestru(self) -> None:
        # 1. Citește perioada selectată
        an_txt = self.combo_an.currentText()
        if not an_txt.isdigit():
            QMessageBox.warning(self, "Eroare", "Selectați un an valid.")
            return
        anul = int(an_txt)
        luni_trimestru = self.trimestre[self.combo_trimestru.currentText()]

        # Inițializează dialog progres
        from utils import ProgressDialog
        progress = ProgressDialog(
            parent=self,
            titlu="Afișare date trimestriale",
            mesaj=f"Se încarcă datele pentru trimestrul selectat..."
        )
        progress.seteaza_valoare(10)

        # 2. Curăță tabelul și lista de date
        self.tabel.setRowCount(0)
        self.date_trimestru.clear()

        try:
            # 3. Interoghează baza DEPCRED
            progress.seteaza_text("Se conectează la baza de date...")
            progress.seteaza_valoare(20)

            conn = sqlite3.connect(DB_DEPCRED)
            cur = conn.cursor()
            placeholders = ",".join(["?"] * len(luni_trimestru))

            progress.seteaza_text("Se execută interogarea pentru datele trimestriale...")
            progress.seteaza_valoare(30)

            cur.execute(
                f"SELECT nr_fisa, luna, dobanda, impr_cred, impr_sold, dep_deb, dep_cred, dep_sold "
                f"FROM depcred WHERE anul=? AND luna IN ({placeholders})",
                [anul] + luni_trimestru
            )
            randuri = cur.fetchall()
            conn.close()

            progress.seteaza_valoare(40)

            if not randuri:
                progress.inchide()
                QMessageBox.information(self, "Info", "Nu există date pentru perioada selectată.")
                self.btn_total.setEnabled(False)
                self.btn_export.setEnabled(False)
                self.btn_export_excel.setEnabled(False)
                return

            # 4. Agregare pe (nr_fisa, luna)
            progress.seteaza_text("Se agregă datele pe fișă și lună...")
            progress.seteaza_valoare(50)

            date_agregate: dict[tuple[int, int], dict[str, float]] = {}
            for nr_fisa, luna, dob, impr_cred, impr_sold, dep_deb, dep_cred, dep_sold in randuri:
                key = (nr_fisa, luna)
                if key not in date_agregate:
                    date_agregate[key] = {
                        'dobanda': 0.0,
                        'impr_cred': 0.0,
                        'dep_deb': 0.0,
                        'dep_cred': 0.0,
                        'impr_sold': impr_sold,
                        'dep_sold': dep_sold
                    }
                date_agregate[key]['dobanda'] += dob or 0.0
                date_agregate[key]['impr_cred'] += impr_cred or 0.0
                date_agregate[key]['dep_deb'] += dep_deb or 0.0
                date_agregate[key]['dep_cred'] += dep_cred or 0.0

            # 5. Colectează intrările și sortează-le după nume
            progress.seteaza_text("Se obțin numele membrilor...")
            progress.seteaza_valoare(65)

            conn_m = sqlite3.connect(DB_MEMBRII)
            cur_m = conn_m.cursor()
            entries = []

            progress.seteaza_text("Se procesează datele...")
            progress.seteaza_valoare(75)

            for (nr_fisa, luna), data in date_agregate.items():
                cur_m.execute("SELECT num_pren FROM membrii WHERE nr_fisa=?", (nr_fisa,))
                row = cur_m.fetchone()
                nume = row[0] if row else "Necunoscut"

                # Calculează total_plata și adaugă în dicționar
                total_plata = data['dobanda'] + data['impr_cred'] + data['dep_deb']
                entries.append((nume, nr_fisa, luna, data, total_plata))
            conn_m.close()

            # Sortare alfabetic (case‑insensitive), apoi după lună
            entries.sort(key=lambda x: (x[0].lower(), x[2]))

            # Populare date_trimestru
            progress.seteaza_text("Se pregătesc datele pentru afișare...")
            progress.seteaza_valoare(85)

            for nume, nr_fisa, luna, data, total_plata in entries:
                self.date_trimestru.append({
                    "nr_fisa": nr_fisa,
                    "luna": luna,
                    "nume": nume,
                    **data,
                    "total_plata": total_plata
                })

            # MODIFICARE - Setăm valorile inițiale de sortare și actualizăm tabelul
            self.sort_column = 2  # Nume
            self.sort_order = Qt.AscendingOrder

            progress.seteaza_text("Se afișează datele în tabel...")
            progress.seteaza_valoare(95)

            self.actualizeaza_tabel()

            progress.seteaza_valoare(100)

            # 8. Activează butoanele Total și Export
            self.btn_total.setEnabled(bool(self.date_trimestru))
            self.btn_export.setEnabled(bool(self.date_trimestru))
            self.btn_export_excel.setEnabled(bool(self.date_trimestru))

        except sqlite3.Error as err:
            progress.inchide()
            QMessageBox.critical(self, "Eroare BD", str(err))
            return
        finally:
            # Închide dialogul de progres
            progress.inchide()

    def afiseaza_totaluri(self) -> None:
        if not self.date_trimestru:
            return
        tot_dep = sum(d["dep_deb"] for d in self.date_trimestru)
        tot_impr = sum(d["impr_cred"] for d in self.date_trimestru)
        tot_dob = sum(d["dobanda"] for d in self.date_trimestru)
        tot_dep_sold = sum(d["dep_sold"] for d in self.date_trimestru)
        tot_impr_sold = sum(d["impr_sold"] for d in self.date_trimestru)
        # Adăugați calcul pentru totaluri noi:
        tot_retragere_fs = sum(d["dep_cred"] for d in self.date_trimestru)
        tot_total_plata = sum(d["dobanda"] + d["impr_cred"] + d["dep_deb"] for d in self.date_trimestru)

        # Actualizați mesajul:
        msg = (
            f"Totaluri {self.combo_trimestru.currentText()} {self.combo_an.currentText()}\n\n"
            f"- Total depuneri (cotizații): {tot_dep:.2f} \n"
            f"- Total rate achitate (împrumuturi): {tot_impr:.2f} \n"
            f"- Total retrageri FS: {tot_retragere_fs:.2f} \n"
            f"- Total dobândă: {tot_dob:.2f} \n"
            f"-------------------------------------------\n"
            f"- Total general plătit: {tot_total_plata:.2f} \n"
            f"-------------------------------------------\n"
            f"- Sold total depuneri: {tot_dep_sold:.2f} \n"
            f"- Sold total împrumuturi: {tot_impr_sold:.2f} \n"
        )
        QMessageBox.information(self, "Totaluri trimestru", msg)

    def _draw_row(self, pdf, col_x, y, row_data, col_widths_mm, is_header=False, bg_color=None):
        """
        Draws a row in the PDF with formatting.

        Args:
            pdf: The reportlab canvas object.
            col_x: List of x-coordinates for column starting positions.
            y: The y-coordinate for the bottom of the row.
            row_data: A list of cell values (strings).
            col_widths_mm: List of column widths in millimeters (before scaling).
            is_header: Boolean indicating if it's the header row.
            bg_color: The background color for the row (reportlab Color object) or None.
        """
        row_h = 8 * mm if not is_header else 20 * mm  # Adjusted header row height to accommodate multiple lines

        for idx, val in enumerate(row_data):
            # Draw background rectangle
            if bg_color:
                pdf.setFillColor(bg_color)
                pdf.rect(col_x[idx], y, col_x[idx + 1] - col_x[idx], row_h, fill=1)

            # Set text color
            pdf.setFillColor(black)
            if not is_header and (
                    (idx == 4 and str(val) == "NEACHITAT") or  # Rata împrumut
                    (idx == 6 and str(val) == "NEACHITAT")  # Depunere lunară
            ):
                pdf.setFillColor(red)

            # Draw text - Handle multiline headers
            if is_header and '\n' in str(val):
                pdf.setFont(DEFAULT_FONT, 10)  # Font for header
                lines = str(val).split('\n')
                line_height = 4 * mm  # Estimated line height, adjust if needed
                # Calculate starting y for the top line to center the block of text vertically within the cell
                total_text_height = len(lines) * line_height
                y_text_start = y + (row_h - total_text_height) / 2 + (
                        line_height * (len(lines) - 1)) - 1 * mm  # Adjusted vertical start slightly

                for i, line in enumerate(lines):
                    pdf.drawString(col_x[idx] + 2 * mm, y_text_start - i * line_height, line)
            else:
                pdf.setFont(DEFAULT_FONT, 9 if not is_header else 10)  # Font for data or single-line header
                # Draw single line text, vertically center if it's a header
                text_y = y + 3 * mm if not is_header else y + row_h / 2 - (
                        4 * mm) / 2 + 1 * mm  # Simplified vertical centering for single line header
                pdf.drawString(col_x[idx] + 2 * mm, text_y, str(val))

    # --- Modified exporta_pdf function ---
    def exporta_pdf(self) -> None:
        if not self.date_trimestru:
            QMessageBox.warning(self, "Lipsă date", "Nu există date de exportat.")
            return

        trimestru_text = self.combo_trimestru.currentText()
        anul = self.combo_an.currentText()
        default_name = f"Situatie_{trimestru_text.replace(' ', '_')}_{anul}.pdf"
        fname, _ = QFileDialog.getSaveFileName(self, "Salvează PDF", default_name, "PDF (*.pdf)")
        if not fname:
            return

        # Inițializare dialog progres
        from utils import ProgressDialog
        progress = ProgressDialog(
            parent=self,
            titlu="Export PDF",
            mesaj="Se generează documentul PDF..."
        )
        progress.seteaza_valoare(10)

        try:
            progress.seteaza_text("Se inițializează documentul PDF...")

            pdf = canvas.Canvas(fname, pagesize=landscape(A4))
            width, height = landscape(A4)
            margin_x, margin_y = 1.5 * cm, 1.5 * cm
            usable_w = width - 2 * margin_x

            # Titlu
            pdf.setFont(DEFAULT_FONT, 14)
            title = f"Situație {trimestru_text} {anul}"
            pdf.drawCentredString(width / 2, height - margin_y, title)

            progress.seteaza_valoare(20)
            progress.seteaza_text("Se configurează structura tabelului...")

            # Coloane - Define widths based on your UI table or desired PDF layout
            headers = [
                "Luna-AN", "Număr\nFișă", "Nume\nPrenume", "Dobândă", "Rată\nÎmprumut",
                "Sold\nÎmprumut", "Depunere\nLunară", "Retragere\nFS", "Sold\nDepunere", "Total\nde Plată"
            ]
            # These widths should ideally match your QTableWidget column widths scaled for PDF
            # You might need to adjust these based on how they look in the PDF
            col_widths_mm = [22, 18, 70, 22, 25, 25, 25, 25, 25, 30]

            # Calculate column x-positions based on usable width and defined widths
            total_col_width_mm = sum(col_widths_mm)
            scale = usable_w / (total_col_width_mm * mm)
            col_x = [margin_x]
            for w in col_widths_mm:
                col_x.append(col_x[-1] + w * mm * scale)
            col_x = col_x   # Nu elimina ultimul element

            y = height - margin_y - 1.5 * cm  # Starting y position for the table

            progress.seteaza_valoare(30)
            progress.seteaza_text("Se desenează antetul tabelului...")

            # Define group background colors using HexColor
            group_colors = [HexColor("#e8f4ff"), HexColor("#fff5e6")]
            header_bg_color = HexColor("#dce8ff")

            # Draw header row
            # Use the original headers list with newlines and increased row height
            header_row_height = 20 * mm  # Define the height for the header row
            self._draw_row(pdf, col_x, y - header_row_height, headers, col_widths_mm, is_header=True,
                           bg_color=header_bg_color)

            y -= (header_row_height + 5 * mm)  # Move y down after header + some space
            row_h = 8 * mm  # Height of data rows

            progress.seteaza_valoare(40)
            progress.seteaza_text("Se desenează datele în tabel...")

            # Determine group background colors (matching your UI logic)
            current_group = 0
            prev_nr = None

            # Sort date_trimestru by name and then month to match UI sorting
            sorted_data = sorted(self.date_trimestru, key=lambda x: (x['nume'].lower(), x['luna']))
            total_rows = len(sorted_data)

            for idx, data in enumerate(sorted_data):
                # Actualizare progres
                current_progress = 40 + int((idx / total_rows) * 50)
                progress.seteaza_valoare(current_progress)

                if idx % 20 == 0:
                    progress.seteaza_text(f"Se procesează rândul {idx + 1} din {total_rows}...")

                # Check for page break
                if y < margin_y + row_h:
                    pdf.showPage()
                    y = height - margin_y - 1.5 * cm
                    # Redraw header on new page
                    self._draw_row(pdf, col_x, y - header_row_height, headers, col_widths_mm, is_header=True,
                                   bg_color=header_bg_color)
                    y -= (header_row_height + 5 * mm)

                # Determine background color for the row based on member group
                nr_curent = data['nr_fisa']
                if prev_nr is not None and nr_curent != prev_nr:
                    current_group = 1 - current_group  # Switch group color
                prev_nr = nr_curent
                bg_color = group_colors[current_group]

                # Prepare row data, applying "NEACHITAT" logic for display
                row_vals = [
                    f"{data['luna']:02d}-{anul}",
                    str(data['nr_fisa']),
                    data['nume'],
                    f"{data['dobanda']:.2f}",
                    "NEACHITAT" if data['impr_sold'] > 0 and data['impr_cred'] == 0 else f"{data['impr_cred']:.2f}",
                    f"{data['impr_sold']:.2f}",
                    "NEACHITAT" if data['dep_sold'] > 0 and data['dep_deb'] == 0 else f"{data['dep_deb']:.2f}",
                    f"{data['dep_cred']:.2f}",
                    f"{data['dep_sold']:.2f}",
                    f"{data['total_plata']:.2f}"
                ]

                # Draw the data row with formatting
                self._draw_row(pdf, col_x, y - row_h, row_vals, col_widths_mm, is_header=False, bg_color=bg_color)

                y -= row_h  # Move to the next row position

            # Add Total row at the end if needed (optional, based on your UI)
            # You would calculate totals here and draw another row similar to the data rows
            # without a background color or with a distinct one.

            progress.seteaza_valoare(95)
            progress.seteaza_text("Se salvează documentul PDF...")

            pdf.save()
            progress.seteaza_valoare(100)

            QMessageBox.information(self, "Export PDF", f"Fișier salvat în: {fname}")
        except Exception as err:
            progress.inchide()
            QMessageBox.critical(self, "Eroare PDF", str(err))
        finally:
            # Închide dialogul de progres
            progress.inchide()

    # MODIFICARE - Adăugare metodă export Excel
    def exporta_excel(self) -> None:
        """
        Exportă datele trimestriale într-un fișier Excel.
        """
        # Verificare date
        if not self.date_trimestru:
            QMessageBox.warning(self, "Lipsă Date", "Nu există date de exportat. Afișați mai întâi trimestrul dorit.")
            return

        # Obține trimestrul și anul din combo boxes
        trimestru_text = self.combo_trimestru.currentText()
        anul = self.combo_an.currentText()

        # Dialog selecție fișier
        default_filename = f"Situatie_{trimestru_text.replace(' ', '_')}_{anul}.xlsx"
        file_name, _ = QFileDialog.getSaveFileName(
            self, "Salvează Excel", default_filename, "Excel Files (*.xlsx)"
        )

        if not file_name:
            return

        # Inițializare dialog progres
        from utils import ProgressDialog
        progress = ProgressDialog(
            parent=self,
            titlu="Export Excel",
            mesaj="Se generează documentul Excel..."
        )
        progress.seteaza_valoare(10)

        try:
            progress.seteaza_text("Se creează workbook-ul Excel...")

            # Creează workbook și sheet
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = f"Situatie_{trimestru_text.replace(' ', '_')}"

            # Definire antet
            progress.seteaza_valoare(20)
            progress.seteaza_text("Se configurează antetul...")

            headers = [
                "Luna-An", "Număr Fișă", "Nume Prenume", "Dobândă", "Rată Împrumut",
                "Sold Împrumut", "Cotizație", "Retragere FS", "Sold Depunere", "Total de Plată"
            ]

            # Definiții stiluri
            header_font = Font(name='Arial', size=11, bold=True)
            header_fill = PatternFill(start_color="DCE8FF", end_color="DCE8FF", fill_type="solid")
            header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

            data_font = Font(name='Arial', size=10)
            data_alignment = Alignment(horizontal='right', vertical='center')
            data_alignment_left = Alignment(horizontal='left', vertical='center')

            # Stiluri pentru grupuri alternante
            group_fill_1 = PatternFill(start_color="E8F4FF", end_color="E8F4FF", fill_type="solid")
            group_fill_2 = PatternFill(start_color="FFF5E6", end_color="FFF5E6", fill_type="solid")

            # Adaugă antet
            for col_idx, header_text in enumerate(headers, 1):
                cell = sheet.cell(row=1, column=col_idx, value=header_text)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment

            # Setează lățimi coloane - ajustate pentru Excel
            progress.seteaza_valoare(30)
            progress.seteaza_text("Se configurează lățimile coloanelor...")

            col_widths = [10, 10, 28, 12, 15, 15, 15, 15, 15, 15]
            for i, width in enumerate(col_widths, 1):
                sheet.column_dimensions[get_column_letter(i)].width = width

            # Adaugă date - sortate după nume și apoi lună
            progress.seteaza_valoare(40)
            progress.seteaza_text("Se adaugă datele în Excel...")

            sorted_data = sorted(self.date_trimestru, key=lambda x: (x['nume'].lower(), x['luna']))

            # Variabile pentru grupuri
            current_group = 0
            prev_nr = None
            total_rows = len(sorted_data)

            for row_idx, data in enumerate(sorted_data, 2):  # Start de la rândul 2 (sub antet)
                # Actualizare progres
                current_progress = 40 + int((row_idx - 2) / total_rows * 50)
                progress.seteaza_valoare(current_progress)

                if row_idx % 20 == 0:
                    progress.seteaza_text(f"Se procesează rândul {row_idx - 1} din {total_rows}...")

                # Alternează grupuri pentru colorare
                nr_curent = data.get('nr_fisa')
                if prev_nr is not None and nr_curent != prev_nr:
                    current_group = 1 - current_group
                prev_nr = nr_curent

                # Alege culoarea de fundal pentru grup
                row_fill = group_fill_1 if current_group % 2 == 0 else group_fill_2

                # Luna-An
                cell = sheet.cell(row=row_idx, column=1, value=f"{data['luna']:02d}-{anul}")
                cell.font = data_font
                cell.alignment = data_alignment
                cell.fill = row_fill

                # Număr Fișă
                cell = sheet.cell(row=row_idx, column=2, value=data.get('nr_fisa', ''))
                cell.font = data_font
                cell.alignment = data_alignment
                cell.fill = row_fill

                # Nume Prenume
                cell = sheet.cell(row=row_idx, column=3, value=data.get('nume', 'Necunoscut'))
                cell.font = data_font
                cell.alignment = data_alignment_left  # Aliniere la stânga pentru nume
                cell.fill = row_fill

                # Dobândă
                cell = sheet.cell(row=row_idx, column=4, value=data.get('dobanda', 0.0))
                cell.font = data_font
                cell.alignment = data_alignment
                cell.fill = row_fill
                cell.number_format = '0.00'

                # Rată Împrumut (cu tratare NEACHITAT)
                if data.get('impr_sold', 0.0) > 0 and data.get('impr_cred', 0.0) == 0:
                    cell = sheet.cell(row=row_idx, column=5, value="NEACHITAT")
                    cell.font = Font(name='Arial', size=10, color="FF0000")  # Red font
                else:
                    cell = sheet.cell(row=row_idx, column=5, value=data.get('impr_cred', 0.0))
                    cell.font = data_font
                    cell.number_format = '0.00'
                cell.alignment = data_alignment
                cell.fill = row_fill

                # Sold Împrumut
                cell = sheet.cell(row=row_idx, column=6, value=data.get('impr_sold', 0.0))
                cell.font = data_font
                cell.alignment = data_alignment
                cell.fill = row_fill
                cell.number_format = '0.00'

                # Cotizație (cu tratare NEACHITAT)
                if data.get('dep_sold', 0.0) > 0 and data.get('dep_deb', 0.0) == 0:
                    cell = sheet.cell(row=row_idx, column=7, value="NEACHITAT")
                    cell.font = Font(name='Arial', size=10, color="FF0000")  # Red font
                else:
                    cell = sheet.cell(row=row_idx, column=7, value=data.get('dep_deb', 0.0))
                    cell.font = data_font
                    cell.number_format = '0.00'
                cell.alignment = data_alignment
                cell.fill = row_fill

                # Retragere FS
                cell = sheet.cell(row=row_idx, column=8, value=data.get('dep_cred', 0.0))
                cell.font = data_font
                cell.alignment = data_alignment
                cell.fill = row_fill
                cell.number_format = '0.00'

                # Sold Depunere
                cell = sheet.cell(row=row_idx, column=9, value=data.get('dep_sold', 0.0))
                cell.font = data_font
                cell.alignment = data_alignment
                cell.fill = row_fill
                cell.number_format = '0.00'

                # Total de Plată
                cell = sheet.cell(row=row_idx, column=10, value=data.get('total_plata', 0.0))
                cell.font = data_font
                cell.alignment = data_alignment
                cell.fill = row_fill
                cell.number_format = '0.00'

            # Adaugă rând de totaluri
            progress.seteaza_valoare(90)
            progress.seteaza_text("Se adaugă totalurile...")

            total_row = len(sorted_data) + 2  # Rândul pentru totaluri

            # Calculează totalurile
            tot_dep = sum(d["dep_deb"] for d in self.date_trimestru)
            tot_impr = sum(d["impr_cred"] for d in self.date_trimestru)
            tot_dob = sum(d["dobanda"] for d in self.date_trimestru)
            tot_dep_sold = sum(d["dep_sold"] for d in self.date_trimestru)
            tot_impr_sold = sum(d["impr_sold"] for d in self.date_trimestru)
            tot_retragere_fs = sum(d["dep_cred"] for d in self.date_trimestru)
            tot_total_plata = sum(d["total_plata"] for d in self.date_trimestru)

            # Stil totaluri
            total_font = Font(name='Arial', size=11, bold=True)
            total_fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")

            # Scrie label pentru totaluri
            cell = sheet.cell(row=total_row, column=1, value="TOTAL:")
            cell.font = total_font
            cell.fill = total_fill

            # Extinde "TOTAL:" pe 3 coloane
            sheet.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=3)

            # Scrie valorile totalurilor
            totals_values = [
                tot_dob, tot_impr, tot_impr_sold,
                tot_dep, tot_retragere_fs, tot_dep_sold, tot_total_plata
            ]

            for col_idx, total_value in enumerate(totals_values, 4):
                cell = sheet.cell(row=total_row, column=col_idx, value=total_value)
                cell.font = total_font
                cell.alignment = data_alignment
                cell.fill = total_fill
                cell.number_format = '0.00'

            # Fixează antetul pentru scroll
            sheet.freeze_panes = "A2"

            # Salvează workbook
            progress.seteaza_valoare(95)
            progress.seteaza_text("Se salvează documentul Excel...")

            workbook.save(file_name)
            progress.seteaza_valoare(100)

            QMessageBox.information(self, "Export reușit", f"Fișierul Excel a fost salvat:\n{file_name}")

        except PermissionError:
            progress.inchide()
            QMessageBox.critical(self, "Eroare Permisiune",
                                 f"Nu s-a putut salva fișierul:\n{file_name}\nVerificați permisiunile sau dacă fișierul este deschis.")
        except Exception as err:
            progress.inchide()
            QMessageBox.critical(self, "Eroare la exportul Excel", f"A apărut o eroare:\n{err}")
        finally:
            # Închide dialogul de progres
            progress.inchide()


if __name__ == "__main__":
    # MODIFICARE - Folosim căile definite
    if not os.path.exists(DB_MEMBRII) or not os.path.exists(DB_DEPCRED):
        QMessageBox.critical(None, "Eroare", "Baze de date MEMBRII.db sau DEPCRED.db lipsă!")
        sys.exit(1)

    app = QApplication(sys.argv)
    if "Fusion" in QtWidgets.QStyleFactory.keys():
        app.setStyle("Fusion")

    wnd = QtWidgets.QMainWindow()
    wnd.setWindowTitle("Vizualizare Situație Trimestrială CAR")
    widget = VizualizareTrimestrialaWidget()
    wnd.setCentralWidget(widget)
    wnd.setMinimumSize(900, 600)
    wnd.show()
    sys.exit(app.exec_())