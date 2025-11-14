import sys
import sqlite3
import os
from datetime import datetime
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP

from PyQt5 import QtWidgets
from PyQt5.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QLabel, QComboBox, QPushButton,
    QTableWidget, QTableWidgetItem, QLineEdit, QMessageBox, QHeaderView,
    QFileDialog, QApplication, QProgressDialog
)
from PyQt5.QtCore import Qt, QTimer, QLocale
from PyQt5.QtGui import QBrush, QColor, QCursor, QDoubleValidator
import openpyxl
from utils import afiseaza_warning, afiseaza_eroare, afiseaza_info, afiseaza_intrebare, ProgressDialog

# Determina calea catre resurse (baze de date) la rulare
if getattr(sys, 'frozen', False):
    BASE_RESOURCE_PATH = os.path.dirname(sys.executable)
else:
    BASE_RESOURCE_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..")

DB_MEMBRII = os.path.join(BASE_RESOURCE_PATH, "MEMBRII.db")
DB_DEPCRED = os.path.join(BASE_RESOURCE_PATH, "DEPCRED.db")
DB_ACTIVI = os.path.join(BASE_RESOURCE_PATH, "ACTIVI.db")
DB_LICHIDATI = os.path.join(BASE_RESOURCE_PATH, "LICHIDATI.db")


class DividendeWidget(QWidget):
    """Widget pentru calculul È™i alocarea dividendelor anuale."""

    def __init__(self, parent=None):
        super().__init__(parent)
        self.an_selectat = datetime.now().year
        # Lista cu membrii È™i dividendele calculate
        self.membri_cu_dividend = []

        self._init_db()
        self._init_ui()
        self._connect_signals()
        self._set_cursors()
        self._load_years()

    def _init_db(self):
        """IniÈ›ializeazÄƒ bazele de date necesare."""
        try:
            conn = None
            try:
                conn = sqlite3.connect(DB_ACTIVI)
                cursor = conn.cursor()
                # Schema oficialÄƒ ACTIVI conform conversie_widget
                cursor.execute("""
                    CREATE TABLE IF NOT EXISTS ACTIVI (
                        NR_FISA INTEGER PRIMARY KEY,
                        NUM_PREN TEXT,
                        DEP_SOLD REAL, -- Soldul din Decembrie
                        DIVIDEND REAL DEFAULT 0.0 -- CÃ¢mpul oficial pentru dividende
                    )
                """)
                # VerificÄƒm È™i adÄƒugÄƒm coloana DIVIDEND dacÄƒ nu existÄƒ
                cursor.execute("PRAGMA table_info(ACTIVI)")
                columns = [col[1] for col in cursor.fetchall()]
                if 'DIVIDEND' not in columns:
                    cursor.execute("ALTER TABLE ACTIVI ADD COLUMN DIVIDEND REAL DEFAULT 0.0")

                conn.commit()
            except sqlite3.Error as e:
                QMessageBox.critical(self, "Eroare BazÄƒ de Date",
                                     f"Eroare la iniÈ›ializarea bazei de date {DB_ACTIVI}: {e}")
                # self.setEnabled(False)
            finally:
                if conn:
                    conn.close()

            if not os.path.exists(DB_MEMBRII):
                QMessageBox.critical(self, "Eroare BazÄƒ de Date",
                                     f"Baza de date esenÈ›ialÄƒ '{DB_MEMBRII}' nu a fost gÄƒsitÄƒ la calea specificatÄƒ: {DB_MEMBRII}")
                # self.setEnabled(False)
            if not os.path.exists(DB_DEPCRED):
                QMessageBox.critical(self, "Eroare BazÄƒ de Date",
                                     f"Baza de date esenÈ›ialÄƒ '{DB_DEPCRED}' nu a fost gÄƒsitÄƒ la calea specificatÄƒ: {DB_DEPCRED}")
                # self.setEnabled(False)

        except Exception as e:
            QMessageBox.critical(self, "Eroare GeneralÄƒ", f"A apÄƒrut o eroare neaÈ™teptatÄƒ la iniÈ›ializarea BD: {e}")
            # self.setEnabled(False)

    def _init_ui(self):
        """IniÈ›ializeazÄƒ interfaÈ›a graficÄƒ."""
        layout = QVBoxLayout(self)

        # Selector An
        an_layout = QHBoxLayout()
        an_layout.addWidget(QLabel("SelectaÈ›i anul pentru calculul dividendelor:"))
        self.combo_an = QComboBox()
        self.combo_an.setMinimumWidth(80)
        an_layout.addWidget(self.combo_an)
        an_layout.addStretch()
        layout.addLayout(an_layout)

        # Inputuri Financiare (acum P - Profit)
        fin_layout = QHBoxLayout()
        fin_layout.addWidget(QLabel("Profit total (P) pentru anul selectat:"))
        self.edit_profit = QLineEdit()  # Am redenumit inputul pentru a fi clar ca e Profitul (P)
        self.edit_profit.setPlaceholderText("0.00")
        validator = QDoubleValidator(-999999999.99, 999999999.99, 2)  # Profitul poate fi si negativ
        validator.setLocale(QLocale(QLocale.C))
        self.edit_profit.setValidator(validator)
        fin_layout.addWidget(self.edit_profit)

        # Am eliminat inputul pentru Cheltuieli totale, deoarece formula foloseste Profitul deja calculat (Venituri - Cheltuieli)
        # fin_layout.addSpacing(20)
        # fin_layout.addWidget(QLabel("Cheltuieli totale (An):"))
        # self.edit_cheltuieli = QLineEdit()
        # self.edit_cheltuieli.setPlaceholderText("0.00")
        # validator_cheltuieli = QDoubleValidator(0.00, 999999999.99, 2)
        # validator_cheltuieli.setLocale(QLocale(QLocale.C))
        # self.edit_cheltuieli.setValidator(validator_cheltuieli)
        # fin_layout.addWidget(self.edit_cheltuieli)

        layout.addLayout(fin_layout)

        # Butoane
        buttons_layout = QHBoxLayout()
        self.btn_clear_activi = QPushButton("È˜terge Date Calcul Anterioare")  # Am schimbat textul sa fie mai clar
        self.btn_populeaza_calculeaza = QPushButton("CalculeazÄƒ Dividende")
        self.btn_transfera_dividend = QPushButton("TransferÄƒ Dividende Ã®n Sold")
        self.btn_transfera_dividend.setEnabled(False)
        self.btn_export_excel = QPushButton("ExportÄƒ Calcul Ã®n Excel")  # Am schimbat textul
        self.btn_export_excel.setEnabled(False)

        buttons_layout.addWidget(self.btn_clear_activi)
        buttons_layout.addWidget(self.btn_populeaza_calculeaza)
        buttons_layout.addWidget(self.btn_transfera_dividend)
        buttons_layout.addWidget(self.btn_export_excel)
        layout.addLayout(buttons_layout)

        # Tabel (coloanele vor afisa datele relevante noii formule)
        self.tabel_dividende = QTableWidget()
        self.tabel_dividende.setColumnCount(5)  # Adaugam o coloana pentru Suma Soldurilor Lunare
        self.tabel_dividende.setHorizontalHeaderLabels([
            "Nr. fiÈ™Äƒ", "Nume È™i prenume", "Sold Dec. An Calcul",  # Pastram soldul din Dec. pentru referinta
            "Suma Solduri Lunare (S membru)",  # Noua coloana
            "Dividend Calculat (B)"
        ])
        # Setam marimea coloanelor
        self.tabel_dividende.setColumnWidth(0, 60)
        self.tabel_dividende.horizontalHeader().setSectionResizeMode(1, QHeaderView.Stretch)
        self.tabel_dividende.setColumnWidth(2, 150)
        self.tabel_dividende.setColumnWidth(3, 250)  # Latime mai mare pentru noua coloana
        self.tabel_dividende.setColumnWidth(4, 150)

        # Setam alinierea implicita a header-elor pe centru
        for i in range(self.tabel_dividende.columnCount()):
            header_item = self.tabel_dividende.horizontalHeaderItem(i)
            if header_item:
                header_item.setTextAlignment(Qt.AlignCenter)

        self.tabel_dividende.setAlternatingRowColors(True)
        layout.addWidget(self.tabel_dividende)

        # Stiluri (pastram stilurile existente)
        self.setStyleSheet("""
            QWidget {font-family:Arial; font-size:10pt; background:#f8f8f8;}
            QLabel {color: #555; margin-right: 5px;}
            QComboBox, QLineEdit {
                background-color: #ffffff; border: 1px solid #b0c0b0;
                border-radius: 5px; padding: 4px; margin: 2px;
            }
             QLineEdit:focus, QComboBox:focus {border-color: #3498db; box-shadow: 0 0 0 2px rgba(52, 152, 219, 0.25); outline: none;}

            QPushButton {
                background:#90EE90; color:#000; border:1px solid #60c060;
                border-radius:5px; padding: 6px 12px; font-weight:bold;
                margin: 5px;
            }
            QPushButton:hover {background:#77dd77;}
            QPushButton:pressed {background:#60c060;}
            QPushButton:disabled {
                background:#d0e0d0;
                color:#808080;
                border:1px solid #b0c0b0;
            }
            QTableWidget {background:#ffffff; alternate-background-color:#f0f0f0; margin-top:10px; border: 1px solid #d0d0d0;}
            QTableWidget::item {padding:4px;}
            QHeaderView::section {
                background-color: #dce8ff; color:#333; padding: 6px;
                border: 1px solid #c0c8d0; font-weight:bold;
                white-space: pre-wrap;
            }
            """)

    def _set_cursors(self):
        """SeteazÄƒ cursorul de mÃ¢nÄƒ pentru elemente interactive."""
        hand_widgets = [
            self.combo_an, self.btn_clear_activi,
            self.btn_populeaza_calculeaza, self.btn_transfera_dividend,  # Am schimbat numele butonului
            self.btn_export_excel
        ]
        for widget in hand_widgets:
            if widget is not None:
                widget.setCursor(QCursor(Qt.PointingHandCursor))

    def _connect_signals(self):
        """ConecteazÄƒ semnalele butoanelor la sloturi."""
        self.combo_an.currentIndexChanged.connect(self._an_selection_changed)
        self.btn_clear_activi.clicked.connect(self._clear_activi)
        self.btn_populeaza_calculeaza.clicked.connect(self._populeaza_activi_calculeaza)
        self.btn_transfera_dividend.clicked.connect(self._transfera_dividend)
        self.btn_export_excel.clicked.connect(self._export_excel)

    def _load_years(self):
        """ÃŽncarcÄƒ ani disponibili Ã®n combobox."""
        years = set()
        conn = None
        try:
            conn = sqlite3.connect(DB_DEPCRED)
            cursor = conn.cursor()
            cursor.execute("SELECT DISTINCT ANUL FROM DEPCRED")
            years.update(row[0] for row in cursor.fetchall())
        except sqlite3.Error as e:
            QMessageBox.warning(self, "Eroare BD", f"Eroare la Ã®ncÄƒrcarea anilor: {e}")
        finally:
            if conn:
                conn.close()

        current_year = datetime.now().year
        for y in range(current_year - 10, current_year + 3):
            years.add(y)

        sorted_years = sorted(list(years), reverse=True)
        self.combo_an.addItems([str(y) for y in sorted_years if y is not None])

        if str(current_year) in sorted_years:
            self.combo_an.setCurrentText(str(current_year))
        elif sorted_years:
            self.combo_an.setCurrentText(str(sorted_years[0]))

        self._an_selection_changed()

    def _an_selection_changed(self):
        """ActualizeazÄƒ anul selectat cÃ¢nd se schimbÄƒ selecÈ›ia Ã®n combobox."""
        try:
            self.an_selectat = int(self.combo_an.currentText())
            self.tabel_dividende.setRowCount(0)
            self.membri_cu_dividend = []  # Am schimbat numele listei interne
            self.edit_profit.clear()  # Acum golim doar inputul de Profit
            # self.edit_cheltuieli.clear() # Eliminat
            self.btn_transfera_dividend.setEnabled(False)  # Am schimbat numele butonului
            self.btn_export_excel.setEnabled(False)

        except ValueError:
            self.an_selectat = None
            QMessageBox.warning(self, "An Invalid", "Anul selectat nu este valid.")
            self.btn_transfera_dividend.setEnabled(False)  # Am schimbat numele butonului
            self.btn_export_excel.setEnabled(False)

    def _clear_activi(self):
        """GoleÈ™te baza de date ACTIVI.db."""
        reply = QMessageBox.question(
            self, "Confirmare",
            "SunteÈ›i sigur cÄƒ doriÈ›i sÄƒ È™tergeÈ›i datele calculate anterior (din tabel È™i baza de date 'Activi')?",
            QMessageBox.Yes | QMessageBox.No
        )
        if reply == QMessageBox.Yes:
            conn = None
            try:
                conn = sqlite3.connect(DB_ACTIVI)
                cursor = conn.cursor()
                cursor.execute("DELETE FROM ACTIVI")
                conn.commit()
                self.tabel_dividende.setRowCount(0)
                self.membri_cu_dividend = []  # Am schimbat numele listei interne
                self.btn_transfera_dividend.setEnabled(False)  # Am schimbat numele butonului
                self.btn_export_excel.setEnabled(False)
                QMessageBox.information(self, "Succes", "Datele calculate anterior au fost golite.")
            except sqlite3.Error as e:
                QMessageBox.critical(self, "Eroare BD", f"Eroare la golirea bazei de date 'Activi': {e}")
            finally:
                if conn:
                    conn.close()

    # ===== METODE DE VALIDARE (FIX 2) =====
    def _validate_member_data(self, cursor_depcred, an_calcul, an_viitor):
        """
        EfectueazÄƒ 3 verificÄƒri de integritate (consistente cu GenerareLuna):
        1. Membri din DEPCRED fÄƒrÄƒ corespondent Ã®n MEMBRII
        2. Membri activi (nu lichidaÈ›i) complet omiÈ™i din anul de calcul
        3. Membri cu DEP_SOLD = 0 sau â‰¤ 0.005 Ã®n Decembrie (ineligibili)

        Pentru cazurile 2 È™i 3, afiÈ™eazÄƒ ultima perioadÄƒ cu sold activ.
        ReturneazÄƒ lista de probleme gÄƒsite sau None dacÄƒ totul este OK.
        """
        issues = []

        # ATTACH LICHIDATI.db pentru a verifica membri lichidaÈ›i
        full_db_lichidati_path = os.path.abspath(DB_LICHIDATI)
        lichidati_attached = False
        try:
            cursor_depcred.execute("ATTACH DATABASE ? AS lichidati_db", (full_db_lichidati_path,))
            lichidati_attached = True
        except sqlite3.Error as e:
            # DacÄƒ LICHIDATI.db nu existÄƒ sau nu poate fi ataÈ™at, continuÄƒm fÄƒrÄƒ (presupunem cÄƒ nu existÄƒ lichidaÈ›i)
            print(f"Avertisment: Nu s-a putut ataÈ™a LICHIDATI.db: {e}")

        # Verificare 1: Membri Ã®n DEPCRED fÄƒrÄƒ corespondent Ã®n MEMBRII
        cursor_depcred.execute("""
            SELECT DISTINCT d.NR_FISA
            FROM DEPCRED d
            LEFT JOIN memb_db.MEMBRII m ON d.NR_FISA = m.NR_FISA
            WHERE d.ANUL = ? AND m.NR_FISA IS NULL
        """, (an_calcul,))

        for (nr_fisa,) in cursor_depcred.fetchall():
            issues.append({
                "nr_fisa": nr_fisa,
                "nume": "NECUNOSCUT",
                "problema": "Membru Ã®n DEPCRED fÄƒrÄƒ Ã®nregistrare Ã®n MEMBRII.db"
            })

        # Verificare 2: Membri activi (nu lichidaÈ›i) care lipsesc complet din anul de calcul
        if lichidati_attached:
            lichidati_check = "AND l.NR_FISA IS NULL"
            lichidati_join = "LEFT JOIN lichidati_db.LICHIDATI l ON m.NR_FISA = l.NR_FISA"
        else:
            lichidati_check = ""
            lichidati_join = ""

        cursor_depcred.execute(f"""
            SELECT
                m.NR_FISA,
                m.NUM_PREN,
                (SELECT MAX(d2.ANUL || '-' || printf('%02d', d2.LUNA))
                 FROM DEPCRED d2
                 WHERE d2.NR_FISA = m.NR_FISA AND d2.DEP_SOLD > 0.005) as ULTIMA_ACTIVITATE
            FROM memb_db.MEMBRII m
            {lichidati_join}
            WHERE m.NR_FISA IS NOT NULL {lichidati_check}
              AND NOT EXISTS (
                  SELECT 1 FROM DEPCRED d
                  WHERE d.NR_FISA = m.NR_FISA AND d.ANUL = ?
              )
        """, (an_calcul,))

        for nr_fisa, nume, ultima_activitate in cursor_depcred.fetchall():
            ultima_act_str = f", ultima activitate: {ultima_activitate}" if ultima_activitate else ", nicio activitate Ã®nregistratÄƒ"
            issues.append({
                "nr_fisa": nr_fisa,
                "nume": nume or "NECUNOSCUT",
                "problema": f"LipseÈ™te complet din {an_calcul}{ultima_act_str}"
            })

        # Verificare 3: Membri cu DEP_SOLD = 0 sau â‰¤ 0.005 Ã®n Decembrie (ineligibili)
        cursor_depcred.execute("""
            SELECT
                d.NR_FISA,
                m.NUM_PREN,
                d.DEP_SOLD,
                (SELECT MAX(d2.ANUL || '-' || printf('%02d', d2.LUNA))
                 FROM DEPCRED d2
                 WHERE d2.NR_FISA = d.NR_FISA AND d2.DEP_SOLD > 0.005) as ULTIMA_ACTIVITATE
            FROM DEPCRED d
            JOIN memb_db.MEMBRII m ON d.NR_FISA = m.NR_FISA
            WHERE d.ANUL = ? AND d.LUNA = 12
              AND (d.DEP_SOLD IS NULL OR d.DEP_SOLD <= 0.005)
        """, (an_calcul,))

        for nr_fisa, nume, dep_sold, ultima_activitate in cursor_depcred.fetchall():
            sold_str = f"{dep_sold:.2f}" if dep_sold is not None else "NULL"
            ultima_act_str = f", ultima activitate: {ultima_activitate}" if ultima_activitate else ""
            issues.append({
                "nr_fisa": nr_fisa,
                "nume": nume or "NECUNOSCUT",
                "problema": f"DEP_SOLD = {sold_str} Ã®n Decembrie {an_calcul} (ineligibil){ultima_act_str}"
            })

        # DETACH LICHIDATI.db
        if lichidati_attached:
            try:
                cursor_depcred.execute("DETACH DATABASE lichidati_db")
            except sqlite3.Error as e:
                print(f"Avertisment: Eroare la detaÈ™area LICHIDATI.db: {e}")

        return issues if issues else None

    def _show_validation_dialog(self, issues):
        """
        AfiÈ™eazÄƒ un dialog cu problemele de validare gÄƒsite.
        Include tabel cu detalii È™i buton de export CSV.
        """
        dialog = QMessageBox(self)
        dialog.setIcon(QMessageBox.Warning)
        dialog.setWindowTitle("Probleme de Integritate Date")
        dialog.setText(f"âš ï¸ Au fost gÄƒsite {len(issues)} probleme care trebuie rezolvate Ã®nainte de calcul:")

        # CreÄƒm un widget custom pentru afiÈ™area detaliilor
        details_widget = QWidget()
        details_layout = QVBoxLayout(details_widget)

        # Tabel cu probleme
        table = QTableWidget()
        table.setRowCount(len(issues))
        table.setColumnCount(3)
        table.setHorizontalHeaderLabels(["Nr. FiÈ™Äƒ", "Nume", "ProblemÄƒ"])
        table.horizontalHeader().setSectionResizeMode(2, QHeaderView.Stretch)
        table.setMinimumWidth(600)
        table.setMinimumHeight(300)

        for row, issue in enumerate(issues):
            table.setItem(row, 0, QTableWidgetItem(str(issue["nr_fisa"])))
            table.setItem(row, 1, QTableWidgetItem(issue["nume"]))
            table.setItem(row, 2, QTableWidgetItem(issue["problema"]))

        details_layout.addWidget(table)

        # Buton export CSV
        export_btn = QPushButton(f"ðŸ’¾ ExportÄƒ ListÄƒ Probleme CSV")
        export_btn.clicked.connect(lambda: self._export_validation_issues_csv(issues))
        details_layout.addWidget(export_btn)

        # AfiÈ™Äƒm dialogul cu widget custom
        dialog.setDetailedText("")  # ActivÄƒm zona de detalii
        dialog.layout().addWidget(details_widget, 1, 0, 1, dialog.layout().columnCount())

        dialog.exec_()

    def _export_validation_issues_csv(self, issues):
        """ExportÄƒ problemele de validare Ã®ntr-un fiÈ™ier CSV."""
        default_filename = f"Membri_Problematici_{self.an_selectat}.csv"
        file_path, _ = QFileDialog.getSaveFileName(
            self, "SalveazÄƒ fiÈ™ier CSV", default_filename, "FiÈ™iere CSV (*.csv);;Toate FiÈ™ierele (*)"
        )

        if not file_path:
            return

        try:
            import csv
            with open(file_path, 'w', newline='', encoding='utf-8-sig') as csvfile:
                writer = csv.writer(csvfile)
                writer.writerow(['Nr. FiÈ™Äƒ', 'Nume', 'ProblemÄƒ'])
                for issue in issues:
                    writer.writerow([issue['nr_fisa'], issue['nume'], issue['problema']])

            QMessageBox.information(self, "Export ReuÈ™it",
                                    f"Lista problemelor a fost exportatÄƒ Ã®n:\n{file_path}")
        except Exception as e:
            QMessageBox.critical(self, "Eroare Export",
                                 f"Eroare la exportul CSV: {e}")

    # --- Metoda de calcul actualizatÄƒ ---
    def _populeaza_activi_calculeaza(self):
        """
        IdentificÄƒ membri cu solduri lunare pozitive Ã®n anul selectat,
        calculeazÄƒ dividendul conform formulei B=(P/Stotal)*Smembru
        È™i populeazÄƒ baza de date ACTIVI È™i tabelul UI.
        """
        if self.an_selectat is None:
            QMessageBox.warning(self, "An Invalid", "SelectaÈ›i un an valid mai Ã®ntÃ¢i.")
            return

        # Prelucrare input Profit (P)
        profit_str = self.edit_profit.text().replace(',', '.')
        try:
            # Profitul poate fi pozitiv sau negativ
            profit_P = Decimal(profit_str) if profit_str else Decimal('0.0')
        except InvalidOperation:
            QMessageBox.warning(self, "Intrare InvalidÄƒ", "Valoare numericÄƒ invalidÄƒ pentru Profit.")
            return

        # --- Implementare Feedback Vizual Start ---
        btn = self.btn_populeaza_calculeaza
        original_text = btn.text()
        btn.setEnabled(False)
        btn.setText("Se calculeazÄƒ...")
        QApplication.processEvents()

        conn_depcred = None
        conn_activi = None

        try:
            conn_depcred = sqlite3.connect(DB_DEPCRED)
            cursor_depcred = conn_depcred.cursor()

            # VerificÄƒm existenÈ›a datelor complete pentru anul selectat (Ian-Dec)
            cursor_depcred.execute("""
                SELECT DISTINCT LUNA FROM DEPCRED WHERE ANUL = ?
            """, (self.an_selectat,))
            luni_existente_an_calcul = set(row[0] for row in cursor_depcred.fetchall())

            if len(luni_existente_an_calcul) < 12:
                QMessageBox.warning(self, "LipsÄƒ Date",
                                    f"Lipsesc date pentru unele luni din anul {self.an_selectat}. "
                                    "Formula necesitÄƒ date complete (Ianuarie-Decembrie)."
                                    "RulaÈ›i procesul de 'Generare LunÄƒ NouÄƒ' pentru lunile lipsÄƒ.")
                return  # Returneaza fara a goli ACTIVI sau procesa mai departe

            # VerificÄƒm existenÈ›a lunii Ianuarie pentru anul urmÄƒtor (necesar pentru transfer)
            an_viitor = self.an_selectat + 1
            cursor_depcred.execute("SELECT COUNT(*) FROM DEPCRED WHERE ANUL = ? AND LUNA = 1", (an_viitor,))
            if cursor_depcred.fetchone()[0] == 0:
                QMessageBox.warning(self, "LipsÄƒ Date",
                                    f"Luna Ianuarie {an_viitor} nu existÄƒ! GeneraÈ›i-o mai Ã®ntÃ¢i pentru a putea transfera dividendul.")
                # Continuam cu calculul si afisarea, dar butonul de transfer va ramane dezactivat (controlat la final)
                # return # Nu mai returnam aici, doar avertizam

            # --- ATTACH DATABASE pentru a accesa MEMBRII.db ---
            full_db_membrii_path = os.path.abspath(DB_MEMBRII)
            try:
                cursor_depcred.execute("ATTACH DATABASE ? AS memb_db", (full_db_membrii_path,))
            except sqlite3.Error as e:
                QMessageBox.critical(self, "Eroare ATTACH BD", f"Eroare la ataÈ™area bazei de date MEMBRII.db: {e}")
                return  # Oprim procesul daca attach esueaza

            # --- Pasul 1: CalculeazÄƒ Suma Soldurilor Lunare (S membru) pentru fiecare membru ---
            # IdentificÄƒm membrii care au avut sold > 0 Ã®n oricare lunÄƒ a anului de calcul
            # È™i le calculÄƒm S_membru (suma soldurilor lunare pe 12 luni)
            # MODIFICAT: Am inclus si numele si soldul din Decembrie direct in aceasta interogare
            cursor_depcred.execute("""
                SELECT
                    d.NR_FISA,
                    m.NUM_PREN,
                    SUM(d.DEP_SOLD) as SUMA_SOLDURI_LUNARE,
                    MAX(CASE WHEN d.LUNA = 12 THEN d.DEP_SOLD ELSE 0 END) as SOLD_DECEMBRIE -- Luam soldul din Decembrie
                FROM DEPCRED d
                JOIN memb_db.MEMBRII m ON d.NR_FISA = m.NR_FISA -- JOIN cu tabela din baza de date atasata
                WHERE d.ANUL = ? AND d.DEP_SOLD > 0 -- Doar inregistrarile cu sold pozitiv din anul selectat
                GROUP BY d.NR_FISA, m.NUM_PREN -- Grupam dupa fisa si nume pentru a calcula suma soldurilor per membru
                HAVING SUM(d.DEP_SOLD) > 0 AND MAX(CASE WHEN d.LUNA = 12 THEN d.DEP_SOLD ELSE 0 END) > 0 -- Sold pozitiv total È˜I Ã®n Decembrie
            """, (self.an_selectat,))

            membri_eligibili_raw = cursor_depcred.fetchall()

            # ===== FIX 2: VALIDARE COMPREHENSIVÄ‚ =====
            # VerificÄƒm integritatea datelor ÃŽNAINTE de calcul
            validation_issues = self._validate_member_data(cursor_depcred, self.an_selectat, an_viitor)

            if validation_issues:
                # AfiÈ™Äƒm dialogul cu problemele gÄƒsite
                self._show_validation_dialog(validation_issues)
                # DETACH Ã®nainte de return
                try:
                    cursor_depcred.execute("DETACH DATABASE memb_db")
                except sqlite3.Error:
                    pass
                return  # BlocÄƒm calculul pÃ¢nÄƒ cÃ¢nd problemele sunt rezolvate

            # --- DETACH DATABASE dupa utilizare ---
            try:
                cursor_depcred.execute("DETACH DATABASE memb_db")
            except sqlite3.Error as e:
                print(
                    f"Avertisment: Eroare la detaÈ™area bazei de date MEMBRII.db: {e}")  # Afisam doar avertisment, nu oprim aplicatia

            if not membri_eligibili_raw:
                QMessageBox.information(self, "InformaÈ›ie",
                                        f"Nu s-au gÄƒsit membri cu solduri lunare pozitive cumulate Ã®n anul {self.an_selectat}.")
                return  # Returneaza fara a goli ACTIVI sau procesa mai departe

            # --- Pasul 2: CalculeazÄƒ S total (Suma Soldurilor Lunare cumulate pentru toÈ›i membrii eligibili) ---
            S_total = Decimal('0.0')
            membri_cu_detalii_calculate = []  # Vom stoca detalii aici pentru a le folosi in calculul final si populare

            for nr_fisa, num_pren, suma_solduri_real, sold_decembrie_real in membri_eligibili_raw:
                suma_solduri_dec = Decimal(str(suma_solduri_real))  # Convertim la Decimal
                sold_decembrie_dec = Decimal(str(sold_decembrie_real))  # Convertim la Decimal
                S_total += suma_solduri_dec

                membri_cu_detalii_calculate.append({
                    "nr_fisa": nr_fisa,
                    "num_pren": num_pren,
                    "dep_sold_dec": sold_decembrie_dec,  # Soldul din Decembrie (Decimal)
                    "suma_solduri_lunare": suma_solduri_dec,
                    "dividend": Decimal('0.0')  # Initializam dividendul, va fi calculat mai jos
                })

            if S_total <= 0:
                QMessageBox.information(self, "InformaÈ›ie",
                                        "Suma totalÄƒ a soldurilor lunare cumulate (S total) este 0 sau negativÄƒ. Nu se poate calcula dividendul.")
                return  # Nu se poate calcula daca S_total e 0 sau negativ

            # --- Pasul 3: CalculeazÄƒ Dividendul (B) pentru fiecare membru ---
            # GoleÈ™te ACTIVI.db Ã®nainte de populare
            conn_activi = sqlite3.connect(DB_ACTIVI)
            cursor_activi = conn_activi.cursor()
            cursor_activi.execute("DELETE FROM ACTIVI")
            conn_activi.commit()

            # Goleste tabelul si lista interna
            self.tabel_dividende.setRowCount(0)
            self.membri_cu_dividend = []  # Resetam lista interna

            # CalculÄƒm È™i populÄƒm
            r = 0  # r count pentru randurile din tabel
            for membru_data in membri_cu_detalii_calculate:  # Iteram peste lista cu datele deja extrase
                nr_fisa = membru_data["nr_fisa"]
                num_pren = membru_data["num_pren"]
                dep_sold_dec_dec = membru_data["dep_sold_dec"]  # Soldul din Decembrie
                S_membru = membru_data["suma_solduri_lunare"]

                # Aplicarea formulei: B = (P / S_total) * S_membru
                # Ne asiguram ca folosim Decimal pentru calcule precise
                dividend_B = (profit_P / S_total * S_membru).quantize(Decimal('0.01'),
                                                                       ROUND_HALF_UP)  # Rotunjim la 2 zecimale

                # Actualizam dividendul in lista interna
                membru_data["dividend"] = dividend_B
                self.membri_cu_dividend.append(membru_data)  # Adaugam in lista finala folosita pentru export/transfer

                # Insereaza Ã®n baza de date ACTIVI
                cursor_activi.execute("""
                     INSERT INTO ACTIVI (NR_FISA, NUM_PREN, DEP_SOLD, DIVIDEND)
                     VALUES (?, ?, ?, ?)
                 """, (
                nr_fisa, num_pren, float(dep_sold_dec_dec), float(dividend_B)))  # Convertim Decimal inapoi la float

                # Adauga rand in tabelul UI
                self.tabel_dividende.insertRow(r)
                item_fisa = QTableWidgetItem(str(nr_fisa))
                item_nume = QTableWidgetItem(num_pren)
                item_sold_dec = QTableWidgetItem(f"{dep_sold_dec_dec:.2f}")  # Afisam soldul din Dec.
                item_suma_solduri = QTableWidgetItem(f"{S_membru:.2f}")  # Afisam suma soldurilor lunare
                item_dividend = QTableWidgetItem(f"{dividend_B:.2f}")  # Afisam dividendul calculat

                # Setam alinierea pentru coloanele numerice la dreapta
                item_fisa.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
                item_sold_dec.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
                item_suma_solduri.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
                item_dividend.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)

                self.tabel_dividende.setItem(r, 0, item_fisa)
                self.tabel_dividende.setItem(r, 1, item_nume)
                self.tabel_dividende.setItem(r, 2, item_sold_dec)
                self.tabel_dividende.setItem(r, 3, item_suma_solduri)
                self.tabel_dividende.setItem(r, 4, item_dividend)

                # Aplica stilul de culoare alternativ pe randuri
                bg_color = QColor("#e8f4ff") if (r % 2 == 0) else QColor("#fff5e6")
                brush = QBrush(bg_color)
                for c in range(self.tabel_dividende.columnCount()):
                    item = self.tabel_dividende.item(r, c)
                    if item:
                        item.setBackground(brush)

                r += 1  # Incrementam contorul de randuri

            conn_activi.commit()  # Commit inserarile in ACTIVI.db

            # Activam butoanele de transfer si export DOAR daca sunt membri cu dividend
            if self.membri_cu_dividend:
                # Verificam din nou daca exista Ianuarie anul urmator inainte de a activa butonul de transfer
                cursor_depcred.execute("SELECT COUNT(*) FROM DEPCRED WHERE ANUL = ? AND LUNA = 1", (an_viitor,))
                if cursor_depcred.fetchone()[0] > 0:
                    self.btn_transfera_dividend.setEnabled(True)
                else:
                    # Daca nu exista Ianuarie anul urmator, avertizam si tinem butonul de transfer dezactivat
                    QMessageBox.warning(self, "AtenÈ›ie",
                                        f"Calculul a fost finalizat, dar luna Ianuarie {an_viitor} nu existÄƒ. Butonul 'TransferÄƒ Dividende Ã®n Sold' rÄƒmÃ¢ne dezactivat pÃ¢nÄƒ la generarea acestei luni.")

                self.btn_export_excel.setEnabled(True)

            QMessageBox.information(
                self, "Calcul Complet",
                f"S-au identificat {len(self.membri_cu_dividend)} membri cu solduri lunare cumulate pozitive.\n"
                f"Suma totalÄƒ a soldurilor lunare (S total): {S_total:.2f} lei.\n"
                f"Dividendul a fost calculat, afiÈ™at Ã®n tabel È™i salvat temporar Ã®n '{DB_ACTIVI}'."
            )


        except sqlite3.Error as e:
            QMessageBox.critical(self, "Eroare BD", f"Eroare la citirea/calculul datelor sau popularea 'Activi': {e}")
            self.btn_transfera_dividend.setEnabled(False)
            self.btn_export_excel.setEnabled(False)

        except Exception as e:
            QMessageBox.critical(self, "Eroare GeneralÄƒ", f"A apÄƒrut o eroare neaÈ™teptatÄƒ: {e}")
            self.btn_transfera_dividend.setEnabled(False)
            self.btn_export_excel.setEnabled(False)

        finally:
            if conn_depcred:
                # ATENTIE: Ne asiguram ca facem DETACH DATABASE indiferent daca a fost eroare sau nu
                try:
                    cursor_depcred.execute("DETACH DATABASE memb_db")
                except sqlite3.Error as e:
                    print(
                        f"Avertisment: Eroare la detaÈ™area bazei de date MEMBRII.db Ã®n blocul finally: {e}")  # Afisam doar avertisment

                conn_depcred.close()

            if conn_activi:
                conn_activi.close()
            # --- Implementare Feedback Vizual Stop ---
            btn.setEnabled(True)
            btn.setText(original_text)
            QApplication.processEvents()

    # --- Metoda de transfer actualizatÄƒ (va transfera DIVIDENDUL) ---
    def _transfera_dividend(self):
        """ActualizeazÄƒ Ã®nregistrÄƒrile din DEPCRED.db cu dividendul calculat."""
        if not self.membri_cu_dividend:  # Am schimbat numele listei interne
            QMessageBox.warning(self, "LipsÄƒ Date",
                                "Nu existÄƒ membri cu dividend calculat pentru transfer.")
            return

        reply = QMessageBox.question(
            self, "Confirmare Transfer",
            f"SunteÈ›i sigur cÄƒ doriÈ›i sÄƒ transferaÈ›i dividendul calculat "
            f"pentru anul {self.an_selectat} Ã®n soldurile din Ianuarie {self.an_selectat + 1}?\n\n"
            "AceastÄƒ operaÈ›iune va actualiza baza de date DEPCRED.db È™i nu poate fi anulatÄƒ uÈ™or."
            "AsiguraÈ›i-vÄƒ cÄƒ aÈ›i rulat deja 'Generare LunÄƒ NouÄƒ' pentru Ianuarie "
            f"È™i cÄƒ datele afiÈ™ate Ã®n tabel sunt corecte.",
            QMessageBox.Yes | QMessageBox.No
        )

        if reply == QMessageBox.Yes:
            # --- IniÈ›ializÄƒm bara de progres ---
            progress_dialog = ProgressDialog(
                parent=self,
                titlu="Transfer Dividende",
                mesaj=f"Se transferÄƒ dividendul pentru {len(self.membri_cu_dividend)} membri...",
                min_val=0,
                max_val=len(self.membri_cu_dividend)
            )

            conn_depcred = None
            conn_activi = None  # Adaugat conexiune la ACTIVI pentru a marca transferat
            try:
                conn_depcred = sqlite3.connect(DB_DEPCRED)
                # Conectam si la ACTIVI pentru a marca membrii transferati (optional, dar util)
                conn_activi = sqlite3.connect(DB_ACTIVI)

                cursor_depcred = conn_depcred.cursor()
                cursor_activi = conn_activi.cursor()

                # Initiem tranzactia in DEPCRED
                conn_depcred.execute("BEGIN TRANSACTION")
                # Putem initia si o tranzactie in ACTIVI daca marcam acolo transferul
                # conn_activi.execute("BEGIN TRANSACTION")

                an_viitor = self.an_selectat + 1
                count_updated = 0
                errored_members = []

                total_membri = len(self.membri_cu_dividend)
                for idx, membru_data in enumerate(self.membri_cu_dividend):
                    # VerificÄƒm dacÄƒ utilizatorul a anulat operaÈ›ia
                    if progress_dialog.este_anulat():
                        conn_depcred.rollback()
                        afiseaza_info("Transferul dividendului a fost anulat de utilizator.", self)
                        return

                    # ActualizÄƒm bara de progres
                    progress_dialog.seteaza_valoare(idx)
                    progress_dialog.seteaza_text(
                        f"Se proceseazÄƒ {idx + 1}/{total_membri}: FiÈ™a {membru_data['nr_fisa']}...")

                    nr_fisa = membru_data["nr_fisa"]
                    dividend = membru_data["dividend"]  # Folosim dividendul calculat

                    try:
                        # Cautam inregistrarea existenta pentru Ianuarie anul urmator
                        cursor_depcred.execute("""
                            SELECT DEP_SOLD, DEP_DEB FROM DEPCRED
                            WHERE NR_FISA = ? AND ANUL = ? AND LUNA = 1
                        """, (nr_fisa, an_viitor))
                        row = cursor_depcred.fetchone()

                        if row:
                            sold_existent_ianuarie = Decimal(str(row[0]))
                            dep_deb_existent = Decimal(str(row[1] or '0.00'))

                            # AdÄƒugÄƒm dividendul la DEP_DEB (nu la DEP_CRED)
                            nou_dep_deb = dep_deb_existent + dividend

                            # RecalculÄƒm soldul adÄƒugÃ¢nd dividendul
                            nou_dep_sold = sold_existent_ianuarie + dividend

                            # ActualizÄƒm Ã®nregistrarea: adÄƒugÄƒm dividendul la DEP_DEB È™i actualizÄƒm soldul
                            cursor_depcred.execute("""
                                 UPDATE DEPCRED
                                 SET DEP_DEB = ?, -- AdÄƒugÄƒm dividendul la suma existentÄƒ pe coloana DEP_DEB
                                     DEP_SOLD = ? -- SetÄƒm noul sold calculat
                                 WHERE NR_FISA = ? AND ANUL = ? AND LUNA = 1
                             """, (float(nou_dep_deb), float(nou_dep_sold), nr_fisa, an_viitor))

                            count_updated += 1

                            # Optional: Marca membrul ca transferat in baza de date ACTIVI
                            # cursor_activi.execute("UPDATE ACTIVI SET TRANSFERRED = 1 WHERE NR_FISA = ?", (nr_fisa,))


                        else:
                            # Aceasta ramura NU ar trebui sa fie atinsa daca validarea initiala functioneaza corect
                            errored_members.append(f"FiÈ™a {nr_fisa} - lipseÈ™te Ã®nregistrarea din Ianuarie {an_viitor}")


                    except sqlite3.Error as e:
                        errored_members.append(f"FiÈ™a {nr_fisa} - eroare BD ({e})")
                        # Nu facem rollback aici, vom decide la final in functie de lista errored_members
                    except Exception as e:
                        errored_members.append(f"FiÈ™a {nr_fisa} - eroare generalÄƒ ({e})")

                # SetÄƒm progresul la maxim pentru a indica finalizarea
                progress_dialog.seteaza_valoare(total_membri)
                progress_dialog.seteaza_text("Finalizare transfer...")

                if not errored_members:
                    conn_depcred.commit()
                    # conn_activi.commit() # Commit si in ACTIVI daca am facut modificari acolo

                    # ÃŽnchidem dialogul de progres Ã®nainte de a afiÈ™a mesajul de informare
                    progress_dialog.inchide()

                    QMessageBox.information(
                        self, "Transfer Complet",
                        f"S-au actualizat cu succes {count_updated} Ã®nregistrÄƒri Ã®n DEPCRED.db "
                        f"pentru luna Ianuarie anul {an_viitor} cu dividendul calculat."
                    )
                    # Golim tabelul si datele temporare dupa transferul reusit
                    self.tabel_dividende.setRowCount(0)
                    self.membri_cu_dividend = []
                    self.btn_transfera_dividend.setEnabled(False)
                    self.btn_export_excel.setEnabled(False)

                else:
                    # Daca au existat erori la actualizari individuale, facem rollback la INTREAGA tranzactie DEPCRED
                    conn_depcred.rollback()
                    # conn_activi.rollback() # Rollback si in ACTIVI daca e cazul

                    # ÃŽnchidem dialogul de progres Ã®nainte de a afiÈ™a mesajul de eroare
                    progress_dialog.inchide()

                    error_msg = "UrmÄƒtorii membri NU au putut fi actualizaÈ›i. OperaÈ›iunea a fost anulatÄƒ (rollback total Ã®n DEPCRED):\n" + "\n".join(
                        errored_members)
                    QMessageBox.critical(self, "Eroare BD ParÈ›ialÄƒ (Transfer)", error_msg)


            except sqlite3.Error as e:
                if conn_depcred:
                    conn_depcred.rollback()
                # if conn_activi:
                # conn_activi.rollback()

                # ÃŽnchidem dialogul de progres Ã®nainte de a afiÈ™a mesajul de eroare
                progress_dialog.inchide()

                QMessageBox.critical(self, "Eroare BD CriticÄƒ (Transfer)",
                                     f"Eroare criticÄƒ la transferul dividendului Ã®n DEPCRED.db: {e}\nOperaÈ›iunea a fost anulatÄƒ (rollback).")

            except Exception as e:
                if conn_depcred:
                    conn_depcred.rollback()
                # if conn_activi:
                # conn_activi.rollback()

                # ÃŽnchidem dialogul de progres Ã®nainte de a afiÈ™a mesajul de eroare
                progress_dialog.inchide()

                QMessageBox.critical(self, "Eroare GeneralÄƒ CriticÄƒ (Transfer)",
                                     f"A apÄƒrut o eroare neaÈ™teptatÄƒ la transfer: {e}\nOperaÈ›iunea a fost anulatÄƒ (rollback).")

            finally:
                # ÃŽnchidem bara de progres È™i eliberÄƒm resursele
                progress_dialog.inchide()

                if conn_depcred:
                    conn_depcred.close()
                if conn_activi:  # Inchidem conexiunea la ACTIVI
                    conn_activi.close()
    # --- Metoda de export actualizatÄƒ (exporta DIVIDENDUL si Suma Soldurilor Lunare) ---
    def _export_excel(self):
        """ExportÄƒ datele din tabel (inclusiv Suma Soldurilor Lunare È™i Dividendul) Ã®n format Excel."""
        if not self.membri_cu_dividend:  # Am schimbat numele listei interne
            QMessageBox.warning(self, "LipsÄƒ Date", "Nu existÄƒ date de exportat. CalculaÈ›i dividendul mai Ã®ntÃ¢i.")
            return

        default_filename = f"Dividende_Anual_{self.an_selectat}.xlsx"
        file_path, _ = QFileDialog.getSaveFileName(
            self, "SalveazÄƒ fiÈ™ier Excel", default_filename, "FiÈ™iere Excel (*.xlsx);;Toate FiÈ™ierele (*)"
        )

        if not file_path:
            return  # Utilizatorul a anulat

        # IniÈ›ializÄƒm bara de progres
        progress_dialog = ProgressDialog(
            parent=self,
            titlu="Export Excel",
            mesaj="Se pregÄƒteÈ™te exportul datelor...",
            min_val=0,
            max_val=100  # Vom ajusta mai tÃ¢rziu Ã®n funcÈ›ie de numÄƒrul de membri
        )

        try:
            # Crearea workbook È™i foaie
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = f"Dividende {self.an_selectat}"

            # Definirea stilurilor pentru Excel
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            from openpyxl.utils import get_column_letter

            # ActualizÄƒm bara de progres
            progress_dialog.seteaza_text("Se definesc stilurile Excel...")
            progress_dialog.seteaza_valoare(10)

            # Stiluri pentru antet
            header_font = Font(name='Arial', size=11, bold=True)
            header_fill = PatternFill(start_color="DCE8FF", end_color="DCE8FF", fill_type="solid")
            header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

            # Stiluri pentru date
            data_font = Font(name='Arial', size=10)
            data_alignment_right = Alignment(horizontal='right', vertical='center')
            data_alignment_left = Alignment(horizontal='left', vertical='center')

            # Stiluri pentru rÃ¢nduri alternante
            row_fill_1 = PatternFill(start_color="E8F4FF", end_color="E8F4FF", fill_type="solid")
            row_fill_2 = PatternFill(start_color="FFF5E6", end_color="FFF5E6", fill_type="solid")

            # ActualizÄƒm bara de progres
            progress_dialog.seteaza_text("Se scriu antetele...")
            progress_dialog.seteaza_valoare(15)

            # Scrie antetele cu formatare
            headers = [self.tabel_dividende.horizontalHeaderItem(i).text() for i in
                       range(self.tabel_dividende.columnCount())]

            for col_idx, header_text in enumerate(headers, 1):
                cell = sheet.cell(row=1, column=col_idx, value=header_text)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment

            # SeteazÄƒ lÄƒÈ›imi optimizate pentru coloane
            col_widths = [10, 30, 15, 25, 20]  # LÄƒÈ›imi iniÈ›iale pentru coloane
            for i, width in enumerate(col_widths, 1):
                sheet.column_dimensions[get_column_letter(i)].width = width

            # ActualizÄƒm bara de progres pentru scrierea datelor
            progress_dialog.seteaza_text("Se scriu datele Ã®n Excel...")
            progress_dialog.seteaza_valoare(20)

            # SetÄƒm maximul pentru bara de progres Ã®n funcÈ›ie de numÄƒrul de membri
            total_membri = len(self.membri_cu_dividend)
            progress_dialog.seteaza_interval(0, 100)  # PÄƒstrÄƒm intervalul 0-100 pentru procente

            # Scrie datele din lista interna cu formatare È™i culori alternante
            current_group = 0
            prev_nr_fisa = None

            for row_idx, membru_data in enumerate(self.membri_cu_dividend, 2):  # ÃŽncepe de la rÃ¢ndul 2 (sub antet)
                # VerificÄƒm dacÄƒ utilizatorul a anulat operaÈ›ia
                if progress_dialog.este_anulat():
                    progress_dialog.inchide()
                    afiseaza_info("Exportul Excel a fost anulat.", self)
                    return

                # CalculÄƒm procentul de progres È™i actualizÄƒm bara
                progress_percent = 20 + int((row_idx - 2) / total_membri * 70)  # 20% - 90%
                progress_dialog.seteaza_valoare(progress_percent)
                progress_dialog.seteaza_text(f"Se scriu datele: {row_idx - 1}/{total_membri} membri...")

                # Alternarea culorilor pentru grupuri de numere de fiÈ™Äƒ
                nr_fisa = membru_data.get("nr_fisa", "")
                if prev_nr_fisa is not None and nr_fisa != prev_nr_fisa:
                    current_group = 1 - current_group  # AlterneazÄƒ grupul
                prev_nr_fisa = nr_fisa

                # Alege culoarea pentru rÃ¢nd
                row_fill = row_fill_1 if current_group % 2 == 0 else row_fill_2

                # Nr. fiÈ™Äƒ
                cell = sheet.cell(row=row_idx, column=1, value=nr_fisa)
                cell.font = data_font
                cell.alignment = data_alignment_right
                cell.fill = row_fill

                # Nume prenume
                cell = sheet.cell(row=row_idx, column=2, value=membru_data.get("num_pren", ""))
                cell.font = data_font
                cell.alignment = data_alignment_left  # Aliniere la stÃ¢nga pentru nume
                cell.fill = row_fill

                # Sold Decembrie
                cell = sheet.cell(row=row_idx, column=3, value=float(membru_data.get("dep_sold_dec", Decimal(0.0))))
                cell.font = data_font
                cell.alignment = data_alignment_right
                cell.fill = row_fill
                cell.number_format = '0.00'

                # Suma Soldurilor Lunare
                cell = sheet.cell(row=row_idx, column=4,
                                  value=float(membru_data.get("suma_solduri_lunare", Decimal(0.0))))
                cell.font = data_font
                cell.alignment = data_alignment_right
                cell.fill = row_fill
                cell.number_format = '0.00'

                # Dividend Calculat
                cell = sheet.cell(row=row_idx, column=5, value=float(membru_data.get("dividend", Decimal(0.0))))
                cell.font = data_font
                cell.alignment = data_alignment_right
                cell.fill = row_fill
                cell.number_format = '0.00'

            # ActualizÄƒm bara de progres
            progress_dialog.seteaza_text("Se calculeazÄƒ totalurile...")
            progress_dialog.seteaza_valoare(90)

            # AdaugÄƒ un rÃ¢nd pentru totaluri
            if self.membri_cu_dividend:
                total_row = len(self.membri_cu_dividend) + 2  # RÃ¢ndul pentru totaluri

                # CalculeazÄƒ totalurile
                total_sold_dec = sum(float(d.get("dep_sold_dec", Decimal(0.0))) for d in self.membri_cu_dividend)
                total_suma_solduri = sum(
                    float(d.get("suma_solduri_lunare", Decimal(0.0))) for d in self.membri_cu_dividend)
                total_dividend = sum(float(d.get("dividend", Decimal(0.0))) for d in self.membri_cu_dividend)

                # Stilul pentru totaluri
                total_font = Font(name='Arial', size=11, bold=True)
                total_fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")

                # Scrie eticheta pentru totaluri
                cell = sheet.cell(row=total_row, column=1, value="TOTAL:")
                cell.font = total_font
                cell.fill = total_fill

                # ÃŽmbinÄƒ celulele pentru etichetÄƒ
                sheet.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=2)

                # Scrie valorile totalurilor
                for col_idx, total_value in enumerate([total_sold_dec, total_suma_solduri, total_dividend], 3):
                    cell = sheet.cell(row=total_row, column=col_idx, value=total_value)
                    cell.font = total_font
                    cell.alignment = data_alignment_right
                    cell.fill = total_fill
                    cell.number_format = '0.00'

            # FixeazÄƒ antetul pentru scroll
            sheet.freeze_panes = "A2"

            # ActualizÄƒm bara de progres
            progress_dialog.seteaza_text("Se salveazÄƒ fiÈ™ierul Excel...")
            progress_dialog.seteaza_valoare(95)

            # SalveazÄƒ workbook-ul
            workbook.save(file_path)

            # FinalizÄƒm bara de progres
            progress_dialog.seteaza_valoare(100)
            progress_dialog.seteaza_text("Export finalizat cu succes!")

            # ÃŽnchidem bara de progres Ã®nainte de a afiÈ™a mesajul de confirmare
            progress_dialog.inchide()

            QMessageBox.information(self, "Export Succes", f"Datele au fost exportate cu succes Ã®n:\n{file_path}")

        except Exception as e:
            # ÃŽnchidem bara de progres Ã®nainte de a afiÈ™a mesajul de eroare
            progress_dialog.inchide()
            QMessageBox.critical(self, "Eroare Export", f"A apÄƒrut o eroare la exportul Excel: {e}")

        finally:
            # Ne asigurÄƒm cÄƒ bara de progres este Ã®nchisÄƒ
            progress_dialog.inchide()


# --- Rularea independentÄƒ a widget-ului pentru testare ---
if __name__ == "__main__":
    app = QApplication(sys.argv)
    if "Fusion" in QtWidgets.QStyleFactory.keys():
        app.setStyle("Fusion")

    main_window = QWidget()
    main_layout = QVBoxLayout(main_window)
    main_layout.addWidget(QLabel("<h3>Modul Calcul Dividende Anual</h3>"))
    dividende_widget = DividendeWidget()
    main_layout.addWidget(dividende_widget)

    main_window.setWindowTitle("Calcul Dividende Anual")
    main_window.setMinimumSize(900, 600)  # Ajustat latimea minima pentru noile coloane
    main_window.show()

    sys.exit(app.exec_())