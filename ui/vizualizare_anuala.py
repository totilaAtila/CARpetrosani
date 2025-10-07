"""
Modul pentru widget‑ul de vizualizare anuală.
Prezentare globală pe un an (fără selector de lună).
"""

# pylint: disable=too-many-lines

import os
import sqlite3
import sys
from datetime import datetime

from PyQt5 import QtWidgets
from PyQt5.QtWidgets import (
    QApplication, QWidget, QLabel, QPushButton, QVBoxLayout, QHBoxLayout,
    QScrollArea, QMessageBox, QTableWidget, QTableWidgetItem, QComboBox,
    QHeaderView, QFileDialog
)
from PyQt5.QtGui import QClipboard, QCursor, QColor, QBrush
from PyQt5.QtCore import Qt, QEvent
from reportlab.lib.pagesizes import A4, landscape
from reportlab.pdfgen import canvas
from reportlab.lib.units import mm, cm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.lib.colors import HexColor, red, black  # Import colors for PDF

# Import pentru export Excel
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from utils import ProgressDialog

if getattr(sys, 'frozen', False):
    # Daca aplicatia este impachetata (ruleaza din executabil)
    # os.path.dirname(sys.executable) va fi directorul care contine executabilul (in onedir mode)
    # Sau directorul temporar in onefile mode (dar onefile e mai complicat pt fisiere externe)
    # Presupunem onedir mode si bazele de date sunt langa exe
    BASE_RESOURCE_PATH = os.path.dirname(sys.executable)
else:
    # Daca ruleaza din scriptul sursa Python (in timpul dezvoltarii)
    # __file__ este calea catre scriptul curent (ex. ui/dividende.py)
    # os.path.dirname(__file__) este directorul scriptului curent (ex. ui)
    # os.path.join(..., "..") urca un nivel (la directorul radacina al proiectului)
    BASE_RESOURCE_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..")


# Definim caile catre bazele de date relative la directorul de resurse
DB_MEMBRII = os.path.join(BASE_RESOURCE_PATH, "MEMBRII.db")
DB_DEPCRED = os.path.join(BASE_RESOURCE_PATH, "DEPCRED.db")
DB_ACTIVI = os.path.join(BASE_RESOURCE_PATH, "ACTIVI.db")
DB_INACTIVI = os.path.join(BASE_RESOURCE_PATH, "INACTIVI.db")
DB_CHITANTE = os.path.join(BASE_RESOURCE_PATH, "CHITANTE.db")
DB_LICHIDATI = os.path.join(BASE_RESOURCE_PATH, "LICHIDATI.db")

# ----------------------------------------------------------------------------
#  Înregistrare fonturi PDF  (DejaVu are suport complet pentru diacritice)
# ----------------------------------------------------------------------------
try:
    # MODIFICARE - Folosim BASE_RESOURCE_PATH în loc de BASE_DIR pentru consistență
    ARIAL = os.path.join(BASE_RESOURCE_PATH, "Arial.ttf")

    # Definim posibile căi pentru fonturile DejaVu
    DEJAVU_FILES = [
        "DejaVuSans.ttf",
        "DejaVuSans-Bold.ttf",
        "DejaVuBoldSans.ttf"
    ]

    # Căutăm fonturile DejaVu în directorul de resurse
    DEJAVU = next((os.path.join(BASE_RESOURCE_PATH, f) for f in DEJAVU_FILES
                  if os.path.exists(os.path.join(BASE_RESOURCE_PATH, f))), None)

    # Înregistrăm fonturile disponibile
    if os.path.exists(ARIAL):
        pdfmetrics.registerFont(TTFont("Arial", ARIAL))
    if DEJAVU and os.path.exists(DEJAVU):
        pdfmetrics.registerFont(TTFont("DejaVu", DEJAVU))

    # Setăm fontul implicit
    DEFAULT_FONT = (
        "DejaVu" if "DejaVu" in pdfmetrics.getRegisteredFontNames() else
        "Arial" if "Arial" in pdfmetrics.getRegisteredFontNames() else
        "Helvetica"
    )
except Exception as err:  # pragma: no cover
    print("Eroare la înregistrarea fonturilor PDF:", err)
    DEFAULT_FONT = "Helvetica"


class VizualizareAnualaWidget(QWidget):
    """Widget pentru vizualizarea situației financiare pe un an."""

    def __init__(self):
        super().__init__()
        self.widgets_cu_cursor_mana: list[QtWidgets.QWidget] = []
        self.date_curente: list[dict[str, float]] = []  # This will store sorted data
        # MODIFICARE - Adăugare atribute pentru sortare
        self.sort_order = Qt.AscendingOrder  # Ordinea sortării: implicit ascendentă
        self.sort_column = 1  # Coloana implicită de sortare (nume prenume)
        self._init_ui()

    # ------------------------------------------------------------------
    # UI
    # ------------------------------------------------------------------
    def _init_ui(self) -> None:
        principal = QVBoxLayout(self)

        # Selector An (centrat) ------------------------------------------------
        selector_layout = QHBoxLayout()
        selector_layout.addStretch(1)  # Adaugă spațiu la stânga
        selector_layout.setSpacing(10)  # Adaugă spațiu între elemente

        an_layout = QHBoxLayout()
        lbl_an = QLabel("Selectare an:")
        an_layout.addWidget(lbl_an)
        an_layout.setSpacing(0)
        self.combo_an = QComboBox()
        cur_year = datetime.now().year
        self.combo_an.addItems([str(y) for y in range(cur_year - 25, cur_year + 5)])
        self.combo_an.setCurrentText(str(cur_year))
        self.combo_an.setMinimumWidth(80)  # 4 cifre vizibile integral
        an_layout.addWidget(self.combo_an)

        self.btn_afiseaza = QPushButton("Afișează anul selectat")
        self.btn_afiseaza.clicked.connect(self.afiseaza_an)
        an_layout.addWidget(self.btn_afiseaza)

        self.btn_total = QPushButton("Afișare total an")
        self.btn_total.setEnabled(False)
        self.btn_total.clicked.connect(self.afiseaza_totaluri)
        an_layout.addWidget(self.btn_total)

        self.btn_export = QPushButton("Exportă PDF")
        self.btn_export.setEnabled(False)
        self.btn_export.clicked.connect(self.exporta_pdf)
        an_layout.addWidget(self.btn_export)

        # MODIFICARE - Adăugare buton export Excel
        self.btn_export_excel = QPushButton("Exportă Excel")
        self.btn_export_excel.setEnabled(False)
        self.btn_export_excel.clicked.connect(self.exporta_excel)
        an_layout.addWidget(self.btn_export_excel)

        selector_layout.addLayout(an_layout)

        selector_layout.addStretch(1)  # Adaugă spațiu la dreapta
        principal.addLayout(selector_layout)

        # Tabel ---------------------------------------------------------------
        self.tabel = QTableWidget()
        # Adjusted column count to match lunar/trimestrial excluding LL-AA
        self.tabel.setColumnCount(9)
        # Adjusted headers to match lunar/trimestrial
        self.tabel.setHorizontalHeaderLabels([
            "Nr. fișă", "Nume\nprenume", "Dobândă", "Rată\nîmprumut",
            "Sold\nîmprumut", "Cotizație", "Retragere\nFS", "Sold\ndepunere", "Total\nde plată"
        ])
        # Adjusted column widths to match lunar/trimestrial approximately
        column_widths = [60, 180, 80, 100, 100, 100, 100, 100, 100]
        for i, w in enumerate(column_widths):
            self.tabel.setColumnWidth(i, w)

        # MODIFICARE - Configurare sortare
        self.tabel.setSortingEnabled(False)  # Dezactivăm sortarea automată pentru a o gestiona manual
        header = self.tabel.horizontalHeader()
        header.sectionClicked.connect(self.on_header_clicked)  # Conectăm semnalul de click pentru sortare

        #  self.tabel.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        # Set specific columns to resize based on content if needed, e.g., Name
        self.tabel.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeToContents)
        # Then set stretch for the rest or specific ones
        #  self.tabel.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)  # Keep Stretch for simplicity initially

        self.tabel.setAlternatingRowColors(True)
        # Moved stylesheet content to _aplica_stiluri for consistency
        # self.tabel.setStyleSheet(...)

        scr = QScrollArea()
        scr.setWidgetResizable(True)
        scr.setWidget(self.tabel)
        principal.addWidget(scr)

        self._aplica_stiluri()  # Call apply styles method

        # cursor de mână
        self.widgets_cu_cursor_mana = [
            self.combo_an, self.btn_afiseaza, self.btn_total, self.btn_export,
            self.btn_export_excel  # MODIFICARE - Adăugat butonul de export Excel
        ]
        for wdg in self.widgets_cu_cursor_mana:
            wdg.installEventFilter(self)

    def _aplica_stiluri(self) -> None:
        """
        Aplică stiluri CSS pentru widget.
        """
        self.setStyleSheet(
            """
            QWidget {font-family:Arial; font-size:10pt; background:#f8f8f8;}
            QPushButton {
                background:#90EE90; color:#000; border:1px solid #60c060;
                border-radius:5px; padding:8px 16px; font-weight:bold;
                margin: 5px;
            }
            QPushButton:hover {background:#77dd77;}
            QPushButton:pressed {background:#60c060;}
            QPushButton:disabled {background:#d0e0d0; color:#808080; border:1px solid #b0c0b0;}
            QHeaderView::section {background-color: #dce8ff; color:#333; padding:6px; border:1px solid #c0c8d0;
            font-weight:bold; white-space: pre-wrap; /* Allow wrapping for headers */}
            QComboBox {
                background-color: #f8f8f8;
                border: 1px solid #b0c0b0;
                border-radius: 5px;
                padding: 8px;
                font-size: 10pt;
                margin: 5px;
            }
            QComboBox:focus {
                border-color: #3498db;
                box-shadow: 0 0 0 2px rgba(52, 152, 219, 0.25);
                outline: none;
            }
            QLabel {
                color: #555;
                margin-right: 5px;
            }
             QTableWidget {background:#ffffff; alternate-background-color:#f0f0f0; margin-top:10px;}
            QTableWidget::item {padding:8px;}
            """
        )

    # ------------------------------------------------------------------
    # Evenimente
    # ------------------------------------------------------------------
    def eventFilter(self, obj, event):  # noqa: N802
        if event.type() == QEvent.Enter and obj in self.widgets_cu_cursor_mana:
            QApplication.setOverrideCursor(Qt.PointingHandCursor)
            return True
        if event.type() == QEvent.Leave and obj in self.widgets_cu_cursor_mana:
            QApplication.restoreOverrideCursor()
            return True
        return super().eventFilter(obj, event)

    # MODIFICARE - Adăugare metode pentru sortare
    def on_header_clicked(self, logical_index):
        """
        Gestionează clicurile pe antetul tabelului pentru a sorta datele.
        """
        if not self.date_curente:
            return  # Nu avem date de sortat

        # Inversează ordinea de sortare dacă se face click pe aceeași coloană
        if self.sort_column == logical_index:
            self.sort_order = Qt.DescendingOrder if self.sort_order == Qt.AscendingOrder else Qt.AscendingOrder
        else:
            self.sort_column = logical_index
            self.sort_order = Qt.AscendingOrder

        # Reafișează tabelul cu date sortate
        self.sorteaza_si_afiseaza_date()

    def sorteaza_si_afiseaza_date(self):
        """
        Sortează datele în funcție de coloana și ordinea selectată și actualizează tabelul.
        """
        # Mapare coloane la cheile din dicționarul de date
        column_to_key = {
            0: 'nr_fisa',
            1: 'nume',
            2: 'dobanda',
            3: 'impr_cred',
            4: 'impr_sold',
            5: 'dep_deb',
            6: 'dep_cred',
            7: 'dep_sold',
            8: 'total_de_plata'
        }

        key = column_to_key.get(self.sort_column)
        if key is None:
            return  # Nu putem sorta după această coloană

        # Sortează datele
        reverse = (self.sort_order == Qt.DescendingOrder)

        # Pentru sortare numerică vs. text
        if key == 'nume':
            self.date_curente.sort(key=lambda x: x[key].lower(), reverse=reverse)
        elif key == 'nr_fisa':
            # Asigură-te că nr_fisa este sortat numeric
            self.date_curente.sort(key=lambda x: int(x[key]) if str(x[key]).isdigit() else 0, reverse=reverse)
        else:
            # Sortare numerică pentru celelalte coloane
            self.date_curente.sort(key=lambda x: float(x.get(key, 0)), reverse=reverse)

        # Actualizează tabelul cu datele sortate
        self.actualizeaza_tabel()

    def actualizeaza_tabel(self):
        """
        Actualizează afișarea tabelului cu datele sortate fără a reinteroga baza de date.
        """
        if not self.date_curente:
            return

        self.tabel.setRowCount(0)
        self.tabel.setRowCount(len(self.date_curente))

        current_group = 0
        prev_nr = None

        for r, data in enumerate(self.date_curente):
            # Populate columns based on the header order and data mapping
            self.tabel.setItem(r, 0, QTableWidgetItem(str(data["nr_fisa"])))  # Nr. fișă
            self.tabel.setItem(r, 1, QTableWidgetItem(data["nume"]))  # Nume prenume
            self.tabel.setItem(r, 2, QTableWidgetItem(f"{data['dobanda']:.2f}"))  # Dobândă (Annual Sum)

            # Rata Imprumut (Annual Sum) - Apply red color if 0 and loan sold > 0
            item_impr = QTableWidgetItem(f"{data['impr_cred']:.2f}")
            if data["impr_cred"] == 0.0 and data["impr_sold"] > 0:
                item_impr.setForeground(QBrush(QColor("red")))
            self.tabel.setItem(r, 3, item_impr)

            self.tabel.setItem(r, 4, QTableWidgetItem(f"{data['impr_sold']:.2f}"))
            # Sold Imprumut (Final Annual Balance)

            # Depunere lunară (Annual Sum) - Apply red color if 0 and deposit sold > 0
            item_dep = QTableWidgetItem(f"{data['dep_deb']:.2f}")
            if data["dep_deb"] == 0.0 and data["dep_sold"] > 0:
                item_dep.setForeground(QBrush(QColor("red")))
            self.tabel.setItem(r, 5, item_dep)

            self.tabel.setItem(r, 6, QTableWidgetItem(f"{data['dep_cred']:.2f}"))
            # Retragere FS (Annual Sum)
            self.tabel.setItem(r, 7, QTableWidgetItem(f"{data['dep_sold']:.2f}"))
            # Sold Depunere (Final Annual Balance)
            self.tabel.setItem(r, 8, QTableWidgetItem(f"{data['total_de_plata']:.2f}"))
            # Total de plată (Annual Sum)

            # Apply row background color based on member group
            nr_curent = data['nr_fisa']
            if prev_nr is not None and nr_curent != prev_nr:
                current_group = 1 - current_group  # Switch group color
            prev_nr = nr_curent
            bg_color = QColor("#e8f4ff") if (current_group % 2 == 0) else QColor("#fff5e6")  # Use QColor for PyQt
            brush = QBrush(bg_color)
            for c in range(self.tabel.columnCount()):
                it = self.tabel.item(r, c)
                if it:
                    it.setBackground(brush)

    # ------------------------------------------------------------------
    # Logica principală
    # ------------------------------------------------------------------
    def afiseaza_an(self) -> None:
        an_txt = self.combo_an.currentText()
        if not an_txt.isdigit():
            QMessageBox.warning(self, "Eroare", "Selectați un an valid.")
            return
        anul = int(an_txt)
        self.tabel.setRowCount(0)
        self.date_curente.clear()

        # Inițializează dialog progres
        from utils import ProgressDialog
        progress = ProgressDialog(
            parent=self,
            titlu="Afișare date anuale",
            mesaj=f"Se încarcă datele pentru anul {anul}..."
        )
        progress.seteaza_valoare(10)  # Inițial 10%

        try:
            # Connect to database
            conn_depcred = sqlite3.connect(DB_DEPCRED)
            cur_depcred = conn_depcred.cursor()

            progress.seteaza_text(f"Se interoghează baza de date pentru anul {anul}...")
            progress.seteaza_valoare(20)

            # Execute the query
            cur_depcred.execute(
                """
                SELECT nr_fisa,
                       SUM(dobanda), SUM(impr_cred), SUM(dep_deb), SUM(dep_cred),
                       impr_sold, dep_sold
                FROM depcred WHERE anul=? GROUP BY nr_fisa
                """,
                (anul,)
            )
            depcred_data = cur_depcred.fetchall()
            conn_depcred.close()

            progress.seteaza_valoare(40)

            if not depcred_data:
                QMessageBox.information(self, "Info", "Nu există date pentru anul selectat.")
                self.btn_total.setEnabled(False)
                self.btn_export.setEnabled(False)
                self.btn_export_excel.setEnabled(False)
                progress.inchide()
                return

            # Fetch names for all members with data
            progress.seteaza_text("Se obțin numele membrilor...")
            progress.seteaza_valoare(60)

            conn_membrii = sqlite3.connect(DB_MEMBRII)
            cur_membrii = conn_membrii.cursor()
            nr_fisa_list = [row[0] for row in depcred_data]
            placeholders = ','.join('?' * len(nr_fisa_list))
            cur_membrii.execute(f"SELECT nr_fisa, num_pren FROM membrii WHERE nr_fisa IN ({placeholders})",
                                nr_fisa_list)
            membrii_data = dict(cur_membrii.fetchall())
            conn_membrii.close()

            progress.seteaza_valoare(80)
            progress.seteaza_text("Se procesează datele...")

            # Process combined data and sort
            combined_data = []
            # Ensure the order of variables here matches the SQL query result order
            for nr, sum_dobanda, sum_impr_cred, sum_dep_deb, sum_dep_cred, final_impr_sold, final_dep_sold in (
                    depcred_data):
                name = membrii_data.get(nr, "Necunoscut")
                total_plata_anuala = (sum_dobanda if sum_dobanda is not None else 0.0) + \
                                     (sum_impr_cred if sum_impr_cred is not None else 0.0) + \
                                     (sum_dep_deb if sum_dep_deb is not None else 0.0)

                combined_data.append({
                    "nr_fisa": nr,
                    "nume": name,
                    "dobanda": sum_dobanda if sum_dobanda is not None else 0.0,
                    "impr_cred": sum_impr_cred if sum_impr_cred is not None else 0.0,
                    "impr_sold": final_impr_sold if final_impr_sold is not None else 0.0,
                    "dep_deb": sum_dep_deb if sum_dep_deb is not None else 0.0,
                    "dep_cred": sum_dep_cred if sum_dep_cred is not None else 0.0,
                    "dep_sold": final_dep_sold if final_dep_sold is not None else 0.0,
                    "total_de_plata": total_plata_anuala
                })

            # MODIFICARE - Setăm datele pentru sortare și afișare
            self.date_curente = combined_data

            # MODIFICARE - Sortare inițială după nume (coloana 1)
            self.sort_column = 1
            self.sort_order = Qt.AscendingOrder

            progress.seteaza_valoare(90)
            progress.seteaza_text("Se afișează datele în tabel...")

            self.sorteaza_si_afiseaza_date()  # Folosim metoda de sortare și afișare

            progress.seteaza_valoare(100)

        except sqlite3.Error as err:
            progress.inchide()
            QMessageBox.critical(self, "Eroare BD", str(err))
            return
        finally:
            # Închide dialogul de progres
            progress.inchide()

        self.btn_total.setEnabled(True)
        self.btn_export.setEnabled(True)
        self.btn_export_excel.setEnabled(True)

    def afiseaza_totaluri(self) -> None:
        if not self.date_curente:
            return
        an_txt = self.combo_an.currentText()
        # Calculate totals from the current displayed data (annual sums/finals)
        total_dobanda = sum(d.get("dobanda", 0.0) for d in self.date_curente)
        total_impr_cred = sum(d.get("impr_cred", 0.0) for d in self.date_curente)
        total_impr_sold = sum(d.get("impr_sold", 0.0) for d in self.date_curente)
        total_dep_deb = sum(d.get("dep_deb", 0.0) for d in self.date_curente)
        total_dep_cred = sum(d.get("dep_cred", 0.0) for d in self.date_curente)
        total_dep_sold = sum(d.get("dep_sold", 0.0) for d in self.date_curente)
        total_de_plata_anuala = sum(d.get("total_de_plata", 0.0) for d in self.date_curente)
        # Total paid in the year (deposits + loan payments + interest)

        msg = (
            f"Totaluri financiare pentru anul {an_txt}\n\n"
            f"- Total Dobândă (anual): {total_dobanda:.2f} \n"
            f"- Total Rate achitate Împrumut (anual): {total_impr_cred:.2f} \n"
            f"- Sold total Împrumut (final an): {total_impr_sold:.2f} \n"
            f"- Total Depuneri (anual): {total_dep_deb:.2f} \n"
            f"- Total Retrageri FS (anual): {total_dep_cred:.2f} \n"
            f"- Sold total Depunere (final an): {total_dep_sold:.2f} \n"
            f"-------------------------------------------\n"
            f"- Total general achitat în an: {total_de_plata_anuala:.2f} \n"
        )

        box = QMessageBox(self)
        box.setWindowTitle("Totaluri anuale")
        box.setTextFormat(Qt.PlainText)
        box.setText(msg)
        box.setIcon(QMessageBox.Information)
        copy_btn = box.addButton("Copiază în clipboard", QMessageBox.ActionRole)
        box.addButton(QMessageBox.Ok)
        box.exec_()

        if box.clickedButton() == copy_btn:
            QApplication.clipboard().setText(msg)
            # mesaj scurt după copiere
            QMessageBox.information(
                self,
                "Copiat",
                f"S‑a copiat în clipboard totalul tranzacțiilor pe anul {an_txt}.",
                QMessageBox.Ok
            )

    # ------------------------------------------------------------------
    # PDF ----------------------------------------------------------------
    # ------------------------------------------------------------------
    # Modified _draw_row signature to accept scaled column widths
    @staticmethod
    def _draw_row(pdf, col_x, col_widths_scaled, y, row_data, is_header=False, bg_color=None):
        """
        Draws a row in the PDF with formatting.

        Args:
            pdf: The reportlab canvas object.
            col_x: List of starting x-coordinates for columns.
            col_widths_scaled: List of scaled widths for columns.
            y: The y-coordinate for the bottom of the row.
            row_data: A list of cell values (strings).
            is_header: Boolean indicating if it's the header row.
            bg_color: The background color for the row (reportlab Color object) or None.
        """
        row_h = 8 * mm if not is_header else 20 * mm  # Adjusted header row height

        for idx, val in enumerate(row_data):
            # Draw background rectangle using start position and scaled width
            if bg_color:
                pdf.setFillColor(bg_color)
                pdf.rect(col_x[idx], y, col_widths_scaled[idx], row_h, fill=1)

            # Set text color
            pdf.setFillColor(black)
            # Apply red color for 0.0 values in specific columns if not header
            # Indices correspond to the new annual header order
            if not is_header:
                try:
                    # Check if the value is a number and is 0.0
                    if float(str(val)) == 0.0:
                        # Color red if the annual total is 0.0 for Rate Imprumut and Depunere lunară
                        if idx in [3, 5]:  # Indices for "Rată\nîmprumut" (3) and "Depunere\nlunară" (5)
                            pdf.setFillColor(red)
                except ValueError:
                    # Not a number, ignore for coloring 0.0
                    pass

            # Draw text - Handle multiline headers
            if is_header and '\n' in str(val):
                pdf.setFont(DEFAULT_FONT, 10)  # Font for header
                lines = str(val).split('\n')
                line_height = 4 * mm  # Estimated line height, adjust if needed
                # Calculate starting y for the top line to center the block of text vertically within the cell
                total_text_height = len(lines) * line_height
                # Position text relative to the column's starting x-coordinate
                y_text_start = y + (row_h - total_text_height) / 2 + (line_height * (len(lines) - 1)) - 1*mm
                # Adjusted vertical start slightly

                for i, line in enumerate(lines):
                    pdf.drawString(col_x[idx] + 2 * mm, y_text_start - i * line_height, line)
            else:
                pdf.setFont(DEFAULT_FONT, 9 if not is_header else 10)  # Font for data or single-line header
                # Draw single line text, vertically center if it's a header
                # Estimated single line text height for 10pt is about 4mm
                # Position text relative to the column's starting x-coordinate
                text_y = y + 3 * mm if not is_header else y + row_h/2 - (4*mm)/2 + 1*mm
                # Simplified vertical centering for single line header
                pdf.drawString(col_x[idx] + 2 * mm, text_y, str(val))

    def exporta_pdf(self) -> None:
        if not self.date_curente:
            QMessageBox.warning(self, "Lipsă Date", "Afișați întâi anul dorit.")
            return
        an_txt = self.combo_an.currentText()
        file_name, _ = QFileDialog.getSaveFileName(
            self, "Salvează PDF", f"Situatie_Anuala_{an_txt}.pdf", "PDF Files (*.pdf)"
        )
        if not file_name:
            return

        # Inițializează dialog progres
        from utils import ProgressDialog
        progress = ProgressDialog(
            parent=self,
            titlu="Export PDF",
            mesaj="Se generează PDF..."
        )
        progress.seteaza_valoare(10)

        try:
            pdf = canvas.Canvas(file_name, pagesize=landscape(A4))
            width, height = landscape(A4)
            ml, mt = 1.5 * cm, 1.5 * cm
            usable_w = width - 2 * ml

            progress.seteaza_text("Se creează antet...")
            progress.seteaza_valoare(20)

            pdf.setFont(DEFAULT_FONT, 14)
            title = f"Situație financiară anul {an_txt}"
            pdf.drawCentredString(width / 2, height - mt, title)

            # Headers for PDF - MUST MATCH QTableWidget HEADERS
            headers = [
                "Nr. fișă", "Nume\nprenume", "Dobândă", "Rată\nîmprumut",
                "Sold\nîmprumut", "Depunere\nlunară", "Retragere\nFS", "Sold\ndepunere", "Total\nde plată"
            ]
            col_widths_mm = [17, 70, 25, 25, 25, 25, 25, 25, 35]

            # Calculate scaled column widths
            total_col_width_mm = sum(col_widths_mm)
            scale = usable_w / (total_col_width_mm * mm)
            col_widths_scaled = [w * mm * scale for w in col_widths_mm]

            # Calculate column starting x-positions
            col_x = [ml]
            current_x = ml
            for w_scaled in col_widths_scaled:
                current_x += w_scaled
                col_x.append(current_x)
            col_x = col_x[:-1]  # Remove the last point which is the right margin

            y = height - mt - 1.5 * cm  # Starting y position for the table

            progress.seteaza_text("Se desenează antetul tabelului...")
            progress.seteaza_valoare(30)

            # Define group background colors using HexColor
            group_colors = [HexColor("#e8f4ff"), HexColor("#fff5e6")]
            header_bg_color = HexColor("#dce8ff")

            # Draw header row using the modified _draw_row
            header_row_height = 20 * mm  # Define header row height
            # Pass col_widths_scaled to _draw_row
            self._draw_row(pdf, col_x, col_widths_scaled, y - header_row_height, headers, is_header=True,
                           bg_color=header_bg_color)

            y -= (header_row_height + 5 * mm)  # Move y down after header + some space
            data_row_height = 8 * mm  # Height for data rows

            progress.seteaza_text("Se desenează datele în PDF...")

            # Set total rows and start progress counter
            total_rows = len(self.date_curente)

            # Iterate through sorted data to draw rows
            current_group = 0
            prev_nr = None

            for idx, data in enumerate(self.date_curente):
                # Update progress bar
                progress_percent = 40 + int((idx / total_rows) * 50)
                progress.seteaza_valoare(progress_percent)

                if idx % 10 == 0:
                    progress.seteaza_text(f"Se procesează rândul {idx + 1}/{total_rows}...")

                # Check for page break
                if y < mt + data_row_height:  # Check against margin and data row height
                    pdf.showPage()
                    y = height - mt - 1.5 * cm  # Reset y position on new page
                    # Redraw header on new page - Pass col_widths_scaled
                    self._draw_row(pdf, col_x, col_widths_scaled, y - header_row_height, headers, is_header=True,
                                   bg_color=header_bg_color)
                    y -= (header_row_height + 5 * mm)

                # Determine background color for the row based on member group
                nr_curent = data.get('nr_fisa')  # Use .get for safety
                if prev_nr is not None and nr_curent != prev_nr:
                    current_group = 1 - current_group  # Switch group color
                prev_nr = nr_curent
                bg_color = group_colors[current_group]

                # Prepare row data - Ensure the order matches the headers list exactly
                row_vals = [
                    str(data.get('nr_fisa', '')),  # Nr. fișă
                    data.get('nume', 'Necunoscut'),  # Nume prenume
                    f"{data.get('dobanda', 0.0):.2f}",  # Dobândă (Annual Sum)
                    f"{data.get('impr_cred', 0.0):.2f}",  # Rată Împrumut (Annual Sum)
                    f"{data.get('impr_sold', 0.0):.2f}",  # Sold Împrumut (Final Annual Balance)
                    f"{data.get('dep_deb', 0.0):.2f}",  # Depunere lunară (Annual Sum)
                    f"{data.get('dep_cred', 0.0):.2f}",  # Retragere FS (Annual Sum)
                    f"{data.get('dep_sold', 0.0):.2f}",  # Sold Depunere (Final Annual Balance)
                    f"{data.get('total_de_plata', 0.0):.2f}"  # Total de plată (Annual Sum)
                ]
                # Draw the data row - Pass col_widths_scaled
                self._draw_row(pdf, col_x, col_widths_scaled, y - data_row_height, row_vals, is_header=False,
                               bg_color=bg_color)
                y -= data_row_height  # Move to the next row position

            progress.seteaza_text("Se finalizează PDF-ul...")
            progress.seteaza_valoare(95)

            pdf.save()
            progress.seteaza_valoare(100)

            QMessageBox.information(self, "Export", "PDF salvat cu succes.")
        except Exception as err:
            QMessageBox.critical(self, "Eroare PDF", str(err))
        finally:
            # Închide dialogul de progres
            progress.inchide()

    # MODIFICARE - Adăugare metodă export Excel
    def exporta_excel(self) -> None:
        """
        Exportă datele anuale într-un fișier Excel.
        """
        # Verificare date
        if not self.date_curente:
            QMessageBox.warning(self, "Lipsă Date", "Nu există date de exportat. Afișați mai întâi anul dorit.")
            return

        # Obține anul din combo box
        an_txt = self.combo_an.currentText()

        # Dialog selecție fișier
        default_filename = f"Situatie_Anuala_{an_txt}.xlsx"
        file_name, _ = QFileDialog.getSaveFileName(
            self, "Salvează Excel", default_filename, "Excel Files (*.xlsx)"
        )

        if not file_name:
            return

        # Inițializare bară progres
        from utils import ProgressDialog
        progress = ProgressDialog(
            parent=self,
            titlu="Export Excel",
            mesaj="Se generează documentul Excel..."
        )
        progress.seteaza_valoare(10)

        try:
            progress.seteaza_text("Se creează workbook-ul Excel...")

            # Creează workbook și sheet
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = f"Situatie_{an_txt}"

            # Definire antet
            progress.seteaza_valoare(20)
            progress.seteaza_text("Se configurează antetul...")

            headers = [
                "Nr. fișă", "Nume prenume", "Dobândă", "Rată împrumut",
                "Sold împrumut", "Cotizație", "Retragere FS", "Sold depunere", "Total de plată"
            ]

            # Definire stiluri
            header_font = Font(name='Arial', size=11, bold=True)
            header_fill = PatternFill(start_color="DCE8FF", end_color="DCE8FF", fill_type="solid")
            header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

            data_font = Font(name='Arial', size=10)
            data_alignment = Alignment(horizontal='right', vertical='center')
            data_alignment_left = Alignment(horizontal='left', vertical='center')

            # Stiluri pentru grupuri alternante
            group_fill_1 = PatternFill(start_color="E8F4FF", end_color="E8F4FF", fill_type="solid")
            group_fill_2 = PatternFill(start_color="FFF5E6", end_color="FFF5E6", fill_type="solid")

            # Adaugă antet
            for col_idx, header_text in enumerate(headers, 1):
                cell = sheet.cell(row=1, column=col_idx, value=header_text)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment

            # Setează lățimi coloane
            progress.seteaza_valoare(30)
            progress.seteaza_text("Se configurează lățimile coloanelor...")

            col_widths = [10, 30, 12, 15, 15, 15, 15, 15, 15]
            for i, width in enumerate(col_widths, 1):
                sheet.column_dimensions[get_column_letter(i)].width = width

            # Adaugă date
            progress.seteaza_valoare(40)
            progress.seteaza_text("Se adaugă datele în Excel...")

            current_group = 0
            prev_nr = None
            total_rows = len(self.date_curente)

            for row_idx, data in enumerate(self.date_curente, 2):
                # Actualizare progres
                if row_idx % 10 == 0 or row_idx == total_rows + 1:
                    current_percent = 40 + int(((row_idx - 2) / total_rows) * 50)
                    progress.seteaza_valoare(current_percent)
                    progress.seteaza_text(f"Se procesează rândul {row_idx - 1} din {total_rows}...")

                # Alternează grupuri
                nr_curent = data.get('nr_fisa')
                if prev_nr is not None and nr_curent != prev_nr:
                    current_group = 1 - current_group
                prev_nr = nr_curent

                # Alege culoarea fundal
                row_fill = group_fill_1 if current_group % 2 == 0 else group_fill_2

                # Nr. fișă
                cell = sheet.cell(row=row_idx, column=1, value=data.get('nr_fisa', ''))
                cell.font = data_font
                cell.alignment = data_alignment
                cell.fill = row_fill

                # Nume prenume
                cell = sheet.cell(row=row_idx, column=2, value=data.get('nume', 'Necunoscut'))
                cell.font = data_font
                cell.alignment = data_alignment_left  # Aliniere la stânga pentru nume
                cell.fill = row_fill

                # Dobândă
                cell = sheet.cell(row=row_idx, column=3, value=data.get('dobanda', 0.0))
                cell.font = data_font
                cell.alignment = data_alignment
                cell.fill = row_fill
                cell.number_format = '0.00'

                # Rată împrumut (cu tratare pentru 0 și sold>0)
                cell = sheet.cell(row=row_idx, column=4, value=data.get('impr_cred', 0.0))
                if data.get('impr_cred', 0.0) == 0.0 and data.get('impr_sold', 0.0) > 0:
                    cell.font = Font(name='Arial', size=10, color="FF0000")  # Red font
                else:
                    cell.font = data_font
                cell.alignment = data_alignment
                cell.fill = row_fill
                cell.number_format = '0.00'

                # Sold împrumut
                cell = sheet.cell(row=row_idx, column=5, value=data.get('impr_sold', 0.0))
                cell.font = data_font
                cell.alignment = data_alignment
                cell.fill = row_fill
                cell.number_format = '0.00'

                # Cotizație (cu tratare pentru 0 și sold>0)
                cell = sheet.cell(row=row_idx, column=6, value=data.get('dep_deb', 0.0))
                if data.get('dep_deb', 0.0) == 0.0 and data.get('dep_sold', 0.0) > 0:
                    cell.font = Font(name='Arial', size=10, color="FF0000")  # Red font
                else:
                    cell.font = data_font
                cell.alignment = data_alignment
                cell.fill = row_fill
                cell.number_format = '0.00'

                # Retragere FS
                cell = sheet.cell(row=row_idx, column=7, value=data.get('dep_cred', 0.0))
                cell.font = data_font
                cell.alignment = data_alignment
                cell.fill = row_fill
                cell.number_format = '0.00'

                # Sold depunere
                cell = sheet.cell(row=row_idx, column=8, value=data.get('dep_sold', 0.0))
                cell.font = data_font
                cell.alignment = data_alignment
                cell.fill = row_fill
                cell.number_format = '0.00'

                # Total de plată
                cell = sheet.cell(row=row_idx, column=9, value=data.get('total_de_plata', 0.0))
                cell.font = data_font
                cell.alignment = data_alignment
                cell.fill = row_fill
                cell.number_format = '0.00'

            # Adaugă rând de totaluri
            progress.seteaza_valoare(90)
            progress.seteaza_text("Se adaugă totalurile...")

            total_row = len(self.date_curente) + 2  # Rândul pentru totaluri

            # Calculează totalurile
            total_dobanda = sum(d.get("dobanda", 0.0) for d in self.date_curente)
            total_impr_cred = sum(d.get("impr_cred", 0.0) for d in self.date_curente)
            total_impr_sold = sum(d.get("impr_sold", 0.0) for d in self.date_curente)
            total_dep_deb = sum(d.get("dep_deb", 0.0) for d in self.date_curente)
            total_dep_cred = sum(d.get("dep_cred", 0.0) for d in self.date_curente)
            total_dep_sold = sum(d.get("dep_sold", 0.0) for d in self.date_curente)
            total_de_plata_anuala = sum(d.get("total_de_plata", 0.0) for d in self.date_curente)

            # Stil totaluri
            total_font = Font(name='Arial', size=11, bold=True)
            total_fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")

            # Scrie label pentru totaluri
            cell = sheet.cell(row=total_row, column=1, value="TOTAL:")
            cell.font = total_font
            cell.fill = total_fill

            # Extinde "TOTAL:" pe 2 coloane
            sheet.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=2)

            # Scrie valorile totalurilor
            totals_values = [
                total_dobanda, total_impr_cred, total_impr_sold,
                total_dep_deb, total_dep_cred, total_dep_sold, total_de_plata_anuala
            ]

            for col_idx, total_value in enumerate(totals_values, 3):
                cell = sheet.cell(row=total_row, column=col_idx, value=total_value)
                cell.font = total_font
                cell.alignment = data_alignment
                cell.fill = total_fill
                cell.number_format = '0.00'

            # Fixează antetul pentru scroll
            sheet.freeze_panes = "A2"

            # Salvează workbook
            progress.seteaza_valoare(95)
            progress.seteaza_text("Se salvează documentul Excel...")

            workbook.save(file_name)
            progress.seteaza_valoare(100)

            QMessageBox.information(self, "Export reușit", f"Fișierul Excel a fost salvat:\n{file_name}")

        except PermissionError:
            progress.inchide()
            QMessageBox.critical(self, "Eroare Permisiune",
                                 f"Nu s-a putut salva fișierul:\n{file_name}\nVerificați permisiunile sau dacă fișierul este deschis.")
        except Exception as err:
            progress.inchide()
            QMessageBox.critical(self, "Eroare la exportul Excel", f"A apărut o eroare:\n{err}")
        finally:
            # Închide dialogul de progres
            progress.inchide()

# ---------------------------------------------------------------------------
if __name__ == "__main__":
    # MODIFICARE - Folosim căile definite pentru a verifica existența bazelor de date
    if not os.path.exists(DB_MEMBRII) or not os.path.exists(DB_DEPCRED):
        QMessageBox.critical(None, "Eroare",
                             "Baze de date MEMBRII.db sau DEPCRED.db lipsă!")
        sys.exit(1)

    app = QApplication(sys.argv)
    if "Fusion" in QtWidgets.QStyleFactory.keys():
        app.setStyle("Fusion")

    wnd = QtWidgets.QMainWindow()
    wnd.setWindowTitle("Vizualizare Situație Anuală CAR")
    widget = VizualizareAnualaWidget()
    wnd.setCentralWidget(widget)
    wnd.setMinimumSize(900, 600)
    wnd.show()
    sys.exit(app.exec_())
