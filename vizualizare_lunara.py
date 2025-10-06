import sys
from PyQt5 import QtWidgets  # Import QtWidgets
from PyQt5.QtWidgets import (
    QWidget, QLabel, QPushButton, QVBoxLayout, QHBoxLayout,
    QScrollArea, QMessageBox, QTableWidget, QTableWidgetItem, QComboBox,
    QHeaderView, QFileDialog, QApplication
)
from PyQt5.QtGui import QClipboard, QCursor, QColor, QBrush
from PyQt5.QtCore import Qt, QEvent
from reportlab.lib.pagesizes import A4, landscape
from reportlab.pdfgen import canvas
from reportlab.lib.units import mm, cm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
import os
import sqlite3
from datetime import datetime
import traceback
# Import Color in addition to HexColor, red, black
from reportlab.lib.colors import HexColor, red, black, Color

# Import pentru export Excel
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from utils import ProgressDialog

if getattr(sys, 'frozen', False):
    # Daca aplicatia este impachetata (ruleaza din executabil)
    # os.path.dirname(sys.executable) va fi directorul care contine executabilul (in onedir mode)
    # Sau directorul temporar in onefile mode (dar onefile e mai complicat pt fisiere externe)
    # Presupunem onedir mode si bazele de date sunt langa exe
    BASE_RESOURCE_PATH = os.path.dirname(sys.executable)
else:
    # Daca ruleaza din scriptul sursa Python (in timpul dezvoltarii)
    # __file__ este calea catre scriptul curent (ex. ui/dividende.py)
    # os.path.dirname(__file__) este directorul scriptului curent (ex. ui)
    # os.path.join(..., "..") urca un nivel (la directorul radacina al proiectului)
    BASE_RESOURCE_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..")

# Definim caile catre bazele de date relative la directorul de resurse
DB_MEMBRII = os.path.join(BASE_RESOURCE_PATH, "MEMBRII.db")
DB_DEPCRED = os.path.join(BASE_RESOURCE_PATH, "DEPCRED.db")
DB_ACTIVI = os.path.join(BASE_RESOURCE_PATH, "ACTIVI.db")
DB_INACTIVI = os.path.join(BASE_RESOURCE_PATH, "INACTIVI.db")
DB_CHITANTE = os.path.join(BASE_RESOURCE_PATH, "CHITANTE.db")
DB_LICHIDATI = os.path.join(BASE_RESOURCE_PATH, "LICHIDATI.db")

# Înregistrare fonturi PDF - MODIFICAT pentru a folosi BASE_RESOURCE_PATH
try:
    # Folosim BASE_RESOURCE_PATH în loc de BASE_DIR pentru a asigura
    # căile corecte atât în modul script cât și după împachetare
    ARIAL = os.path.join(BASE_RESOURCE_PATH, "Arial.ttf")

    # Definim posibile căi pentru fonturile DejaVu
    DEJAVU_FILES = [
        "DejaVuSans.ttf",
        "DejaVuSans-Bold.ttf",
        "DejaVuBoldSans.ttf"
    ]

    # Căutăm fonturile DejaVu în directorul de resurse
    DEJAVU = next((os.path.join(BASE_RESOURCE_PATH, f) for f in DEJAVU_FILES
                   if os.path.exists(os.path.join(BASE_RESOURCE_PATH, f))), None)

    # Înregistrăm fonturile disponibile
    if os.path.exists(ARIAL):
        pdfmetrics.registerFont(TTFont("Arial", ARIAL))
    if DEJAVU and os.path.exists(DEJAVU):
        pdfmetrics.registerFont(TTFont("DejaVu", DEJAVU))

    # Setăm fontul implicit
    DEFAULT_FONT = (
        "DejaVu" if "DejaVu" in pdfmetrics.getRegisteredFontNames() else
        "Arial" if "Arial" in pdfmetrics.getRegisteredFontNames() else
        "Helvetica"
    )
except Exception as e:
    print(f"Eroare la înregistrarea fonturilor PDF: {e}")
    DEFAULT_FONT = 'Helvetica'


class VizualizareLunaraWidget(QWidget):
    """
    Widget pentru vizualizarea datelor lunare din DEPCRED.db și MEMBRII.db.
    Utilizează container QWidget pentru centrare și stil CSS standardizat.
    """

    def __init__(self):
        super().__init__()
        self.widgets_cu_cursor_mana = []
        self.date_curente = []  # This was for unsorted data, might not be strictly necessary now
        self.date_sortate_afisate = []  # This will store sorted data for UI and PDF
        self.sort_order = Qt.AscendingOrder  # Ordinea sortării: implicit ascendentă
        self.sort_column = 2  # Coloana implicită de sortare (nume prenume)
        self.init_ui()

    def init_ui(self):
        """
        Inițializează interfața utilizatorului.
        Folosește QWidget container pentru centrarea robustă a rândului de sus.
        """
        layout_principal = QVBoxLayout(self)

        # --- Structura cu QWidget Container pentru CENTRARE ROBUSTĂ ---

        # Layout-ul principal pentru rândul de sus (va centra containerul)
        selector_layout = QHBoxLayout()
        # Nu setăm spațiere aici direct, ci în layout-ul interior dacă e necesar

        # Layout pentru a grupa widget-urile de selecție și butoanele
        labels_layout = QHBoxLayout()
        labels_layout.setSpacing(5)  # Spațiere între controalele din rând

        self.label_luna = QLabel("Luna:")
        labels_layout.addWidget(self.label_luna)
        self.combo_luna = QComboBox()
        # Nu mai setăm MinimumWidth aici, lăsăm stilul CSS sau Qt să decidă
        luni = ["Ianuarie", "Februarie", "Martie", "Aprilie", "Mai", "Iunie",
                "Iulie", "August", "Septembrie", "Octombrie", "Noiembrie", "Decembrie"]
        self.combo_luna.addItems(luni)
        self.combo_luna.setCurrentIndex(datetime.now().month - 1)
        labels_layout.addWidget(self.combo_luna)

        self.label_an = QLabel("An:")
        labels_layout.addWidget(self.label_an)
        self.combo_an = QComboBox()
        # Nu mai setăm MinimumWidth aici
        current_year = datetime.now().year
        years = [str(y) for y in range(current_year - 25, current_year + 5)]
        self.combo_an.addItems(years)
        self.combo_an.setCurrentText(str(current_year))
        labels_layout.addWidget(self.combo_an)

        self.buton_afiseaza = QPushButton("Afișează")
        self.buton_afiseaza.clicked.connect(self.afiseaza_luna)
        labels_layout.addWidget(self.buton_afiseaza)

        self.buton_total = QPushButton("Afișare total luna")
        self.buton_total.setEnabled(False)
        self.buton_total.clicked.connect(self.afiseaza_totaluri)
        labels_layout.addWidget(self.buton_total)

        self.buton_export_pdf = QPushButton("Exportă PDF")
        self.buton_export_pdf.setEnabled(False)
        self.buton_export_pdf.clicked.connect(self.exporta_pdf)
        labels_layout.addWidget(self.buton_export_pdf)

        # MODIFICARE - Adăugare buton export Excel
        self.buton_export_excel = QPushButton("Exportă Excel")
        self.buton_export_excel.setEnabled(False)
        self.buton_export_excel.clicked.connect(self.exporta_excel)
        labels_layout.addWidget(self.buton_export_excel)

        # Creăm widget-ul container intermediar
        container_widget = QWidget()
        # Setăm labels_layout ca layout pentru acest container
        container_widget.setLayout(labels_layout)

        # Adăugăm stretch, widget-ul container, apoi stretch în selector_layout
        selector_layout.addStretch(1)
        selector_layout.addWidget(container_widget)  # Adăugăm containerul
        selector_layout.addStretch(1)

        # --- Sfârșit structură QWidget Container ---

        # Actualizare lista de widgeturi pentru schimbarea cursorului
        self.widgets_cu_cursor_mana = [
            self.buton_afiseaza, self.buton_total, self.buton_export_pdf,
            self.buton_export_excel, self.combo_an, self.combo_luna  # MODIFICARE - Adăugat buton_export_excel
        ]
        for widget in self.widgets_cu_cursor_mana:
            widget.installEventFilter(self)

        # Configurarea tabelului
        self.tabel = QTableWidget()
        self.tabel.setColumnCount(10)
        self.tabel.setHorizontalHeaderLabels([
            "LL-AA", "Nr. fișă", "Nume\nprenume", "Dobândă", "Rată\nîmprumut",
            "Sold\nîmprumut", "Cotizație", "Retragere\nFS", "Sold\ndepunere", "Total\nde plată"
        ])
        column_widths = [60, 60, 200, 80, 100, 100, 100, 100, 100, 120]
        for i, w in enumerate(column_widths):
            self.tabel.setColumnWidth(i, w)

        # Configurare sortare
        self.tabel.setSortingEnabled(False)  # Dezactivăm sortarea automată pentru a o gestiona manual
        header = self.tabel.horizontalHeader()
        header.sectionClicked.connect(self.on_header_clicked)  # Conectăm semnalul de click pentru sortare

        # Stilul specific tabelului va fi setat în aplica_stiluri

        # Add alternating row colors for the table
        self.tabel.setAlternatingRowColors(True)

        # Adăugarea tabelului într-o zonă derulabilă
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setWidget(self.tabel)

        # Adăugarea layout-urilor la layout-ul principal al ferestrei
        layout_principal.addLayout(selector_layout)  # Adăugăm rândul de sus (modificat)
        layout_principal.addWidget(scroll)  # Adăugăm zona cu tabelul
        self.setLayout(layout_principal)
        self.aplica_stiluri()  # Aplicăm stilul standardizat

    def afiseaza_luna(self):
        luna = self.combo_luna.currentIndex() + 1
        try:
            anul = int(self.combo_an.currentText())
        except ValueError:
            QMessageBox.warning(self, "Eroare", "Selectați un an valid din listă.")
            return

        self.tabel.setRowCount(0)
        self.date_curente.clear()
        self.date_sortate_afisate.clear()

        # Inițializare dialog progres
        from utils import ProgressDialog
        progress = ProgressDialog(
            parent=self,
            titlu="Încărcare date lunare",
            mesaj=f"Se încarcă datele pentru luna {luna}/{anul}..."
        )
        progress.seteaza_valoare(10)

        try:
            progress.seteaza_text("Se conectează la baza de date...")
            progress.seteaza_valoare(20)

            conn = sqlite3.connect(DB_DEPCRED)
            conn.execute(f"ATTACH DATABASE '{DB_MEMBRII}' AS membrii_db")
            cursor = conn.cursor()

            progress.seteaza_text("Se execută interogarea...")
            progress.seteaza_valoare(40)

            cursor.execute(
                """
                SELECT d.nr_fisa, m.num_pren, d.dobanda, d.impr_cred, d.impr_sold,
                       d.dep_deb, d.dep_cred, d.dep_sold
                FROM depcred d
                LEFT JOIN membrii_db.membrii m ON d.nr_fisa = m.nr_fisa
                WHERE d.luna = ? AND d.anul = ?
                ORDER BY m.num_pren
                """,
                (luna, anul),
            )
            rows = cursor.fetchall()
            conn.close()

            progress.seteaza_valoare(60)

            if not rows:
                progress.inchide()
                QMessageBox.information(self, "Info", "Nu există date pentru luna selectată.")
                self.buton_total.setEnabled(False)
                self.buton_export_pdf.setEnabled(False)
                self.buton_export_excel.setEnabled(False)
                return

            progress.seteaza_text("Se procesează datele...")
            progress.seteaza_valoare(70)

            # Store fetched data for potential sorting and processing
            fetched_data = []
            for row in rows:
                nr_fisa, nume, dobanda, impr_cred, impr_sold, dep_deb, dep_cred_fetched, dep_sold = row
                fetched_data.append({
                    "nr_fisa": nr_fisa,
                    "nume": nume or "Nume negăsit",
                    "dobanda": dobanda or 0.0,
                    "impr_cred": impr_cred or 0.0,
                    "impr_sold": impr_sold or 0.0,
                    "dep_deb": dep_deb or 0.0,
                    "dep_cred": dep_cred_fetched or 0.0,
                    "dep_sold": dep_sold or 0.0,
                })

            # Sort the data alphabetically by name (case-insensitive)
            self.date_sortate_afisate = sorted(fetched_data, key=lambda x: x['nume'].lower())

            progress.seteaza_text("Se calculează totalurile...")
            progress.seteaza_valoare(80)

            # Calculează total_plata pentru fiecare rând înaintea sortării
            for r, data in enumerate(self.date_sortate_afisate):
                total_plata = data.get('dobanda', 0.0) + data.get('impr_cred', 0.0) + data.get('dep_deb', 0.0)
                self.date_sortate_afisate[r]["total_plata"] = total_plata

            # Sortare inițială după nume (coloana 2)
            self.sort_column = 2
            self.sort_order = Qt.AscendingOrder

            progress.seteaza_text("Se afișează datele în tabel...")
            progress.seteaza_valoare(90)

            # Afișează datele în tabel
            self.actualizeaza_tabel()

            progress.seteaza_valoare(100)

            self.buton_total.setEnabled(bool(self.date_sortate_afisate))
            self.buton_export_pdf.setEnabled(bool(self.date_sortate_afisate))
            self.buton_export_excel.setEnabled(bool(self.date_sortate_afisate))

        except Exception as e:
            progress.inchide()
            QMessageBox.critical(self, "Eroare la interogare", str(e))
            return
        finally:
            # Închide dialogul de progres
            progress.inchide()

    def afiseaza_totaluri(self):
        # Calculate totals from the sorted data
        if not self.date_sortate_afisate:
            QMessageBox.information(self, "Totaluri", "Nu există date afișate pentru a calcula totalurile.")
            return

        # Calculate totals from the sorted data stored in self.date_sortate_afisate
        total_dobanda = sum(d.get("dobanda", 0.0) for d in self.date_sortate_afisate)
        total_impr_cred = sum(d.get("impr_cred", 0.0) for d in self.date_sortate_afisate)
        total_impr_sold = sum(d.get("impr_sold", 0.0) for d in self.date_sortate_afisate)
        total_dep_deb = sum(d.get("dep_deb", 0.0) for d in self.date_sortate_afisate)
        total_dep_cred = sum(d.get("dep_cred", 0.0) for d in self.date_sortate_afisate)  # Corrected to dep_cred
        total_dep_sold = sum(d.get("dep_sold", 0.0) for d in self.date_sortate_afisate)
        total_general_plata = sum(
            d.get("total_plata", 0.0) for d in self.date_sortate_afisate)  # Use the calculated total_plata

        luna_text = self.combo_luna.currentText()
        anul = self.combo_an.currentText()

        mesaj = (
            f"Totaluri financiare pentru {luna_text} {anul}\n\n"
            f"- Total dobândă: {total_dobanda:.2f} \n"
            f"- Total rate achitate (împrumuturi): {total_impr_cred:.2f} \n"
            f"- Sold total împrumut: {total_impr_sold:.2f} \n"
            f"- Total depuneri (cotizații): {total_dep_deb:.2f} \n"
            f"- Total retrageri FS: {total_dep_cred:.2f} \n"  # Corrected to total_dep_cred
            f"- Sold total depuneri: {total_dep_sold:.2f} \n"
            f"-------------------------------------------\n"
            f"- Total general plătit: {total_general_plata:.2f} \n"
        )

        box = QMessageBox(self)
        box.setWindowTitle("Totaluri lunare")
        box.setTextFormat(Qt.PlainText)
        box.setText(mesaj)
        box.setIcon(QMessageBox.Information)
        copy_button = box.addButton("Copiază în clipboard", QMessageBox.ActionRole)
        box.addButton(QMessageBox.Ok)
        box.exec_()

        if box.clickedButton() == copy_button:
            QApplication.clipboard().setText(mesaj)

    # --- Modified _draw_row function for PDF with background color ---
    # Modified signature to accept scaled column widths
    def _draw_row(self, pdf, col_x, col_widths_scaled, y, row_data, is_header=False, bg_color=None):
        """
        Draws a row in the PDF with formatting and optional background color.

        Args:
            pdf: The reportlab canvas object.
            col_x: List of starting x-coordinates for columns.
            col_widths_scaled: List of scaled widths for columns.
            y: The y-coordinate for the bottom of the row.
            row_data: A list of cell values (strings).
            is_header: Boolean indicating if it's the header row.
            bg_color: The background color for the row (reportlab Color object) or None.
        """
        row_h = 8 * mm if not is_header else 20 * mm  # Adjusted header row height

        for idx, val in enumerate(row_data):
            # Draw background rectangle using start position and scaled width
            if bg_color:
                pdf.setFillColor(bg_color)
                pdf.rect(col_x[idx], y, col_widths_scaled[idx], row_h, fill=1)

            # Set text color
            pdf.setFillColor(black)
            # Apply red color for "NEACHITAT" values in specific columns if not header
            # Indices correspond to the lunar header order
            if not is_header:
                if (idx == 4 and str(val) == "NEACHITAT") or \
                        (idx == 6 and str(val) == "NEACHITAT"):
                    pdf.setFillColor(red)

            # Draw text - Handle multiline headers
            if is_header and '\n' in str(val):
                pdf.setFont(DEFAULT_FONT, 10)  # Font for header
                lines = str(val).split('\n')
                line_height = 4 * mm  # Estimated line height, adjust if needed
                # Calculate starting y for the top line to center the block of text vertically within the cell
                total_text_height = len(lines) * line_height
                # Position text relative to the column's starting x-coordinate
                y_text_start = y + (row_h - total_text_height) / 2 + (
                            line_height * (len(lines) - 1)) - 1 * mm  # Adjusted vertical start slightly

                for i, line in enumerate(lines):
                    pdf.drawString(col_x[idx] + 2 * mm, y_text_start - i * line_height, line)
            else:
                pdf.setFont(DEFAULT_FONT, 9 if not is_header else 10)  # Font for data or single-line header
                # Draw single line text, vertically center if it's a header
                # Estimated single line text height for 10pt is about 4mm
                # Position text relative to the column's starting x-coordinate
                text_y = y + 3 * mm if not is_header else y + row_h / 2 - (
                            4 * mm) / 2 + 1 * mm  # Simplified vertical centering for single line header
                pdf.drawString(col_x[idx] + 2 * mm, text_y, str(val))

    def exporta_pdf(self):
        """
        Exportă datele lunare într-un fișier PDF.
        """
        # Use the sorted data for PDF export
        if not self.date_sortate_afisate:
            QMessageBox.warning(self, "Lipsă Date", "Nu există date de exportat. Afișați mai întâi luna dorită.")
            return

        # Get luna and anul from the combo boxes
        luna = self.combo_luna.currentIndex() + 1
        try:
            anul = int(self.combo_an.currentText())
        except ValueError:
            QMessageBox.warning(self, "Eroare", "Selectați un an valid din listă.")
            return

        luna_text = self.combo_luna.currentText()
        default_filename = f"Situatie_Lunara_{luna_text}_{anul}.pdf"
        file_name, _ = QFileDialog.getSaveFileName(
            self, "Salvează PDF", default_filename, "PDF Files (*.pdf)"
        )

        if not file_name:
            return

        # Inițializare dialog progres
        from utils import ProgressDialog
        progress = ProgressDialog(
            parent=self,
            titlu="Export PDF",
            mesaj="Se generează documentul PDF..."
        )
        progress.seteaza_valoare(10)

        try:
            progress.seteaza_text("Se inițializează documentul PDF...")

            pdf = canvas.Canvas(file_name, pagesize=landscape(A4))
            width, height = landscape(A4)
            margin_x, margin_y = 1.5 * cm, 1.5 * cm
            usable_width = width - 2 * margin_x

            # Desenare titlu
            progress.seteaza_valoare(20)
            progress.seteaza_text("Se desenează titlul...")

            pdf.setFont(DEFAULT_FONT, 14)
            title_text = f"Situație financiară lunară - {luna_text} {anul}"
            pdf.drawCentredString(width / 2.0, height - margin_y, title_text)

            # Definire coloane și antet
            progress.seteaza_valoare(30)
            progress.seteaza_text("Se definesc coloanele și antetul...")

            headers = [
                "LL-AA", "Nr. fișă", "Nume\nprenume", "Dobândă", "Rată\nîmprumut",
                "Sold\nîmprumut", "Depunere\nlunară", "Retragere\nFS", "Sold\ndepunere", "Total\nde plată"
            ]
            col_widths_mm = [18, 18, 68, 20, 22, 22, 22, 22, 22, 30]

            # Calculate scaled column widths
            total_col_width_mm = sum(col_widths_mm)
            scale = usable_width / (total_col_width_mm * mm)
            col_widths_scaled = [w * mm * scale for w in col_widths_mm]

            # Calculate column starting x-positions
            col_x = []
            current_x = margin_x
            for w_scaled in col_widths_scaled:
                col_x.append(current_x)
                current_x += w_scaled

            # Define colors
            header_bg_color = HexColor("#dce8ff")
            group_colors = [HexColor("#e8f4ff"), HexColor("#fff5e6")]

            # Calculate position for header
            header_row_height = 20 * mm
            data_row_height = 8 * mm
            y_table_top = height - margin_y - 1.5 * cm
            y_header_bottom = y_table_top - header_row_height
            y_first_row_top = y_header_bottom - 5 * mm
            y = y_first_row_top - data_row_height

            # Draw header
            progress.seteaza_valoare(40)
            progress.seteaza_text("Se desenează antetul tabelului...")

            self._draw_row(pdf, col_x, col_widths_scaled, y_header_bottom, headers, is_header=True,
                           bg_color=header_bg_color)

            # Desenare rânduri date
            progress.seteaza_valoare(50)
            progress.seteaza_text("Se desenează datele în tabel...")

            current_group = 0
            prev_nr = None
            total_rows = len(self.date_sortate_afisate)

            for idx, row_data in enumerate(self.date_sortate_afisate):
                # Actualizare progres
                current_progress = 50 + int((idx / total_rows) * 45)
                progress.seteaza_valoare(current_progress)

                if idx % 20 == 0:
                    progress.seteaza_text(f"Se procesează rândul {idx + 1} din {total_rows}...")

                # Verificare paginare nouă
                if y < margin_y:
                    pdf.showPage()
                    y_new_page_table_top = height - margin_y - 1.5 * cm
                    y_new_page_header_bottom = y_new_page_table_top - header_row_height
                    y_new_page_first_row_top = y_new_page_header_bottom - 5 * mm
                    y = y_new_page_first_row_top - data_row_height

                    # Redesenează antetul pe pagina nouă
                    self._draw_row(pdf, col_x, col_widths_scaled, y_new_page_header_bottom, headers, is_header=True,
                                   bg_color=header_bg_color)

                # Determină culoarea fundal pentru rând
                nr_curent = row_data.get('nr_fisa')
                if prev_nr is not None and nr_curent != prev_nr:
                    current_group = 1 - current_group
                prev_nr = nr_curent
                bg_color = group_colors[current_group]

                # Pregătire date pentru rând
                valori = [
                    f"{luna:02d}-{anul}",
                    str(row_data.get('nr_fisa', '')),
                    row_data.get('nume', 'Nume negăsit'),
                    f"{row_data.get('dobanda', 0.0):.2f}",
                    "NEACHITAT" if row_data.get('impr_sold', 0.0) > 0 and row_data.get('impr_cred',
                                                                                       0.0) == 0 else f"{row_data.get('impr_cred', 0.0):.2f}",
                    f"{row_data.get('impr_sold', 0.0):.2f}",
                    "NEACHITAT" if row_data.get('dep_sold', 0.0) > 0 and row_data.get('dep_deb',
                                                                                      0.0) == 0 else f"{row_data.get('dep_deb', 0.0):.2f}",
                    f"{row_data.get('dep_cred', 0.0):.2f}",
                    f"{row_data.get('dep_sold', 0.0):.2f}",
                    f"{row_data.get('total_plata', 0.0):.2f}"
                ]

                # Desenează rândul cu formatare
                self._draw_row(pdf, col_x, col_widths_scaled, y, valori, is_header=False, bg_color=bg_color)

                # Decrementează poziția y pentru următorul rând
                y -= data_row_height

            # Finalizare PDF
            progress.seteaza_valoare(95)
            progress.seteaza_text("Se salvează documentul PDF...")

            pdf.save()
            progress.seteaza_valoare(100)

            QMessageBox.information(self, "Export reușit", f"Fișierul PDF a fost salvat:\n{file_name}")
        except PermissionError:
            progress.inchide()
            QMessageBox.critical(self, "Eroare Permisiune",
                                 f"Nu s-a putut salva fișierul:\n{file_name}\nVerificați permisiunile sau dacă fișierul este deschis.")
        except Exception as e:
            progress.inchide()
            QMessageBox.critical(self, "Eroare la exportul PDF", f"A apărut o eroare:\n{e}\n\n{traceback.format_exc()}")
        finally:
            # Închide dialogul de progres
            progress.inchide()

    # MODIFICARE - Adăugare metodă exporta_excel
    def exporta_excel(self):
        """
        Exportă datele lunare într-un fișier Excel.
        """
        # Verificare date
        if not self.date_sortate_afisate:
            QMessageBox.warning(self, "Lipsă Date", "Nu există date de exportat. Afișați mai întâi luna dorită.")
            return

        # Obține luna și anul din combo boxes
        luna = self.combo_luna.currentIndex() + 1
        luna_text = self.combo_luna.currentText()
        try:
            anul = int(self.combo_an.currentText())
        except ValueError:
            QMessageBox.warning(self, "Eroare", "Selectați un an valid din listă.")
            return

        # Dialog selecție fișier
        default_filename = f"Situatie_Lunara_{luna_text}_{anul}.xlsx"
        file_name, _ = QFileDialog.getSaveFileName(
            self, "Salvează Excel", default_filename, "Excel Files (*.xlsx)"
        )

        if not file_name:
            return

        # Inițializare dialog progres
        from utils import ProgressDialog
        progress = ProgressDialog(
            parent=self,
            titlu="Export Excel",
            mesaj="Se generează documentul Excel..."
        )
        progress.seteaza_valoare(10)

        try:
            progress.seteaza_text("Se creează workbook-ul Excel...")

            # Creează workbook și sheet
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = f"Situatie_{luna_text}_{anul}"

            # Definire antet
            progress.seteaza_valoare(20)
            progress.seteaza_text("Se configurează antetul...")

            headers = [
                "LL-AA", "Nr. fișă", "Nume prenume", "Dobândă", "Rată împrumut",
                "Sold împrumut", "Cotizație", "Retragere FS", "Sold depunere", "Total de plată"
            ]

            # Definiții stiluri
            header_font = Font(name='Arial', size=11, bold=True)
            header_fill = PatternFill(start_color="DCE8FF", end_color="DCE8FF", fill_type="solid")
            header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

            data_font = Font(name='Arial', size=10)
            data_alignment = Alignment(horizontal='right', vertical='center')
            data_alignment_left = Alignment(horizontal='left', vertical='center')

            # Stiluri pentru grupuri alternante
            group_fill_1 = PatternFill(start_color="E8F4FF", end_color="E8F4FF", fill_type="solid")
            group_fill_2 = PatternFill(start_color="FFF5E6", end_color="FFF5E6", fill_type="solid")

            # Adaugă antet
            for col_idx, header_text in enumerate(headers, 1):
                cell = sheet.cell(row=1, column=col_idx, value=header_text)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment

            # Setează lățimi coloane
            progress.seteaza_valoare(30)
            progress.seteaza_text("Se configurează lățimile coloanelor...")

            col_widths = [10, 10, 28, 12, 15, 15, 15, 15, 15, 15]
            for i, width in enumerate(col_widths, 1):
                sheet.column_dimensions[get_column_letter(i)].width = width

            # Adaugă date
            progress.seteaza_valoare(40)
            progress.seteaza_text("Se adaugă datele în Excel...")

            current_group = 0
            prev_nr = None
            total_rows = len(self.date_sortate_afisate)

            for row_idx, data in enumerate(self.date_sortate_afisate, 2):
                # Actualizare progres
                current_progress = 40 + int((row_idx - 2) / total_rows * 50)
                progress.seteaza_valoare(current_progress)

                if row_idx % 20 == 0:
                    progress.seteaza_text(f"Se procesează rândul {row_idx - 1} din {total_rows}...")

                # Alternează grupuri pentru colorare
                nr_curent = data.get('nr_fisa')
                if prev_nr is not None and nr_curent != prev_nr:
                    current_group = 1 - current_group
                prev_nr = nr_curent

                # Alege culoarea de fundal pentru grup
                row_fill = group_fill_1 if current_group % 2 == 0 else group_fill_2

                # Luna-AN
                cell = sheet.cell(row=row_idx, column=1, value=f"{luna:02d}-{anul}")
                cell.font = data_font
                cell.alignment = data_alignment
                cell.fill = row_fill

                # Nr. fișă
                cell = sheet.cell(row=row_idx, column=2, value=data.get('nr_fisa', ''))
                cell.font = data_font
                cell.alignment = data_alignment
                cell.fill = row_fill

                # Nume prenume
                cell = sheet.cell(row=row_idx, column=3, value=data.get('nume', 'Necunoscut'))
                cell.font = data_font
                cell.alignment = data_alignment_left
                cell.fill = row_fill

                # Dobândă
                cell = sheet.cell(row=row_idx, column=4, value=data.get('dobanda', 0.0))
                cell.font = data_font
                cell.alignment = data_alignment
                cell.fill = row_fill
                cell.number_format = '0.00'

                # Rată împrumut (cu tratare NEACHITAT)
                if data.get('impr_sold', 0.0) > 0 and data.get('impr_cred', 0.0) == 0:
                    cell = sheet.cell(row=row_idx, column=5, value="NEACHITAT")
                    cell.font = Font(name='Arial', size=10, color="FF0000")
                else:
                    cell = sheet.cell(row=row_idx, column=5, value=data.get('impr_cred', 0.0))
                    cell.font = data_font
                    cell.number_format = '0.00'
                cell.alignment = data_alignment
                cell.fill = row_fill

                # Sold împrumut
                cell = sheet.cell(row=row_idx, column=6, value=data.get('impr_sold', 0.0))
                cell.font = data_font
                cell.alignment = data_alignment
                cell.fill = row_fill
                cell.number_format = '0.00'

                # Cotizație (cu tratare NEACHITAT)
                if data.get('dep_sold', 0.0) > 0 and data.get('dep_deb', 0.0) == 0:
                    cell = sheet.cell(row=row_idx, column=7, value="NEACHITAT")
                    cell.font = Font(name='Arial', size=10, color="FF0000")
                else:
                    cell = sheet.cell(row=row_idx, column=7, value=data.get('dep_deb', 0.0))
                    cell.font = data_font
                    cell.number_format = '0.00'
                cell.alignment = data_alignment
                cell.fill = row_fill

                # Retragere FS
                cell = sheet.cell(row=row_idx, column=8, value=data.get('dep_cred', 0.0))
                cell.font = data_font
                cell.alignment = data_alignment
                cell.fill = row_fill
                cell.number_format = '0.00'

                # Sold depunere
                cell = sheet.cell(row=row_idx, column=9, value=data.get('dep_sold', 0.0))
                cell.font = data_font
                cell.alignment = data_alignment
                cell.fill = row_fill
                cell.number_format = '0.00'

                # Total de plată
                total_plata = data.get('total_plata', 0.0)
                cell = sheet.cell(row=row_idx, column=10, value=total_plata)
                cell.font = data_font
                cell.alignment = data_alignment
                cell.fill = row_fill
                cell.number_format = '0.00'

            # Adaugă rând de totaluri
            progress.seteaza_valoare(90)
            progress.seteaza_text("Se adaugă totalurile...")

            total_row = len(self.date_sortate_afisate) + 2

            # Calculează totalurile
            total_dobanda = sum(d.get("dobanda", 0.0) for d in self.date_sortate_afisate)
            total_impr_cred = sum(d.get("impr_cred", 0.0) for d in self.date_sortate_afisate)
            total_impr_sold = sum(d.get("impr_sold", 0.0) for d in self.date_sortate_afisate)
            total_dep_deb = sum(d.get("dep_deb", 0.0) for d in self.date_sortate_afisate)
            total_dep_cred = sum(d.get("dep_cred", 0.0) for d in self.date_sortate_afisate)
            total_dep_sold = sum(d.get("dep_sold", 0.0) for d in self.date_sortate_afisate)
            total_general_plata = sum(d.get("total_plata", 0.0) for d in self.date_sortate_afisate)

            # Stil totaluri
            total_font = Font(name='Arial', size=11, bold=True)
            total_fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")

            # Scrie label pentru totaluri
            cell = sheet.cell(row=total_row, column=1, value="TOTAL:")
            cell.font = total_font
            cell.fill = total_fill

            # Extinde "TOTAL:" pe 3 coloane
            sheet.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=3)

            # Scrie valorile totalurilor
            totals_values = [
                total_dobanda, total_impr_cred, total_impr_sold,
                total_dep_deb, total_dep_cred, total_dep_sold, total_general_plata
            ]

            for col_idx, total_value in enumerate(totals_values, 4):
                cell = sheet.cell(row=total_row, column=col_idx, value=total_value)
                cell.font = total_font
                cell.alignment = data_alignment
                cell.fill = total_fill
                cell.number_format = '0.00'

            # Fixează antetul pentru scroll
            sheet.freeze_panes = "A2"

            # Salvează workbook
            progress.seteaza_valoare(95)
            progress.seteaza_text("Se salvează documentul Excel...")

            workbook.save(file_name)
            progress.seteaza_valoare(100)

            QMessageBox.information(self, "Export reușit", f"Fișierul Excel a fost salvat:\n{file_name}")

        except PermissionError:
            progress.inchide()
            QMessageBox.critical(self, "Eroare Permisiune",
                                 f"Nu s-a putut salva fișierul:\n{file_name}\nVerificați permisiunile sau dacă fișierul este deschis.")
        except Exception as e:
            progress.inchide()
            QMessageBox.critical(self, "Eroare la exportul Excel",
                                 f"A apărut o eroare:\n{e}\n\n{traceback.format_exc()}")
        finally:
            # Închide dialogul de progres
            progress.inchide()

    def aplica_stiluri(self):
        """
        Aplică stilul CSS standardizat, similar cu vizualizare_anuala/trimestriala.
        """
        # Stil preluat din vizualizare_trimestriala.py pentru consistență
        self.setStyleSheet(
            """
            QWidget {font-family:Arial; font-size:10pt; background:#f8f8f8;}
            QPushButton {
                background:#90EE90; color:#000; border:1px solid #60c060;
                border-radius:5px; padding:8px 16px; font-weight:bold;
                margin: 5px;
            }
            QPushButton:hover {background:#77dd77;}
            QPushButton:pressed {background:#60c060;}
            QPushButton:disabled {background:#d0e0d0; color:#808080; border:1px solid #b0c0b0;}
            QHeaderView::section {
                background-color:#dce8ff; color:#333; padding:6px;
                border:1px solid #c0c8d0; font-weight:bold;
                white-space: pre-wrap; /* Allow wrapping for headers */
            }
            QComboBox {
                background-color:#f8f8f8; border:1px solid #b0c0b0;
                border-radius:5px; padding:8px; font-size:10pt; margin:5px;
            }
            QComboBox:focus {
                border-color:#3498db;
                box-shadow: 0 0 0 2px rgba(52, 152, 219, 0.25); /* Păstrăm efectul focus */
                outline:none;
            }
            QLabel {
                color:#555; margin-right:5px;
                /* Am eliminat padding-top specific pentru a se alinia cu stilul standard */
            }
            QTableWidget {
                background:#ffffff; alternate-background-color:#f0f0f0;
                margin-top:10px;
                /* Am eliminat gridline-color, border-bottom, etc. specifice din versiunea anterioară */
                /* pentru a se potrivi cu stilul mai simplu din anuala/trimestriala */
            }
            QTableWidget::item {
                padding:8px; /* Padding standard */
            }
            /* Am eliminat QTableWidget::item:selected specific */
            /* Am eliminat QComboBox::drop-down specific */
            /* Am eliminat QComboBox QAbstractItemView specific */
            /* Am eliminat QScrollArea specific */
            """
        )

    def on_header_clicked(self, logical_index):
        """
        Gestionează clicurile pe antetul tabelului pentru a sorta datele.
        """
        if not self.date_sortate_afisate:
            return  # Nu avem date de sortat

        # Inversează ordinea de sortare dacă se face click pe aceeași coloană
        if self.sort_column == logical_index:
            self.sort_order = Qt.DescendingOrder if self.sort_order == Qt.AscendingOrder else Qt.AscendingOrder
        else:
            self.sort_column = logical_index
            self.sort_order = Qt.AscendingOrder

        # Reafișează tabelul cu date sortate
        self.sorteaza_si_afiseaza_date()

    def sorteaza_si_afiseaza_date(self):
        """
        Sortează datele în funcție de coloana și ordinea selectată și actualizează tabelul.
        """
        # Mapare coloane la cheile din dicționarul de date
        column_to_key = {
            0: None,  # LL-AA (nu avem nevoie de sortare)
            1: 'nr_fisa',
            2: 'nume',
            3: 'dobanda',
            4: 'impr_cred',
            5: 'impr_sold',
            6: 'dep_deb',
            7: 'dep_cred',
            8: 'dep_sold',
            9: 'total_plata'
        }

        key = column_to_key.get(self.sort_column)
        if key is None:
            return  # Nu putem sorta după această coloană

        # Sortează datele
        reverse = (self.sort_order == Qt.DescendingOrder)

        # Pentru sortare numerică vs. text
        if key == 'nume':
            self.date_sortate_afisate.sort(key=lambda x: x[key].lower(), reverse=reverse)
        elif key == 'nr_fisa':
            # Asigură-te că nr_fisa este sortat numeric
            self.date_sortate_afisate.sort(key=lambda x: int(x[key]) if str(x[key]).isdigit() else 0, reverse=reverse)
        else:
            # Sortare numerică pentru celelalte coloane
            self.date_sortate_afisate.sort(key=lambda x: float(x.get(key, 0)), reverse=reverse)

        # Actualizează tabelul cu datele sortate
        self.actualizeaza_tabel()

    def actualizeaza_tabel(self):
        """
        Actualizează afișarea tabelului cu datele sortate fără a reinteroga baza de date.
        """
        luna = self.combo_luna.currentIndex() + 1
        try:
            anul = int(self.combo_an.currentText())
        except ValueError:
            return

        self.tabel.setRowCount(0)
        self.tabel.setRowCount(len(self.date_sortate_afisate))

        current_group = 0
        prev_nr = None

        for r, data in enumerate(self.date_sortate_afisate):
            # Folosește total_plata deja calculat
            total_plata = data.get('total_plata', 0.0)

            # Apply row background color based on member group
            nr_curent = data.get('nr_fisa')
            if prev_nr is not None and nr_curent != prev_nr:
                current_group = 1 - current_group
            prev_nr = nr_curent
            bg_color = QColor("#e8f4ff") if (current_group % 2 == 0) else QColor("#fff5e6")
            brush = QBrush(bg_color)

            # Populate table items and apply background color
            self.tabel.setItem(r, 0, QTableWidgetItem(f"{luna:02d}-{anul}"))
            self.tabel.item(r, 0).setBackground(brush)

            self.tabel.setItem(r, 1, QTableWidgetItem(str(data.get("nr_fisa", ''))))
            self.tabel.item(r, 1).setBackground(brush)

            self.tabel.setItem(r, 2, QTableWidgetItem(data.get("nume", 'Nume negăsit')))
            self.tabel.item(r, 2).setBackground(brush)

            self.tabel.setItem(r, 3, QTableWidgetItem(f"{data.get('dobanda', 0.0):.2f}"))
            self.tabel.item(r, 3).setBackground(brush)

            item_impr = QTableWidgetItem("NEACHITAT" if data.get('impr_sold', 0.0) > 0 and data.get('impr_cred',
                                                                                                    0.0) == 0 else f"{data.get('impr_cred', 0.0):.2f}")
            if item_impr.text() == "NEACHITAT":
                item_impr.setForeground(QBrush(QColor("red")))
            self.tabel.setItem(r, 4, item_impr)
            self.tabel.item(r, 4).setBackground(brush)

            self.tabel.setItem(r, 5, QTableWidgetItem(f"{data.get('impr_sold', 0.0):.2f}"))
            self.tabel.item(r, 5).setBackground(brush)

            item_dep = QTableWidgetItem("NEACHITAT" if data.get('dep_sold', 0.0) > 0 and data.get('dep_deb',
                                                                                                  0.0) == 0 else f"{data.get('dep_deb', 0.0):.2f}")
            if item_dep.text() == "NEACHITAT":
                item_dep.setForeground(QBrush(QColor("red")))
            self.tabel.setItem(r, 6, item_dep)
            self.tabel.item(r, 6).setBackground(brush)

            self.tabel.setItem(r, 7, QTableWidgetItem(f"{data.get('dep_cred', 0.0):.2f}"))
            self.tabel.item(r, 7).setBackground(brush)

            self.tabel.setItem(r, 8, QTableWidgetItem(f"{data.get('dep_sold', 0.0):.2f}"))
            self.tabel.item(r, 8).setBackground(brush)

            self.tabel.setItem(r, 9, QTableWidgetItem(f"{total_plata:.2f}"))
            self.tabel.item(r, 9).setBackground(brush)

    def eventFilter(self, obj, event):
        """
        Filtrează evenimentele pentru a schimba cursorul mouse-ului.
        """
        if event.type() == QEvent.Enter and obj in self.widgets_cu_cursor_mana:
            QApplication.setOverrideCursor(Qt.PointingHandCursor)
            return True
        elif event.type() == QEvent.Leave and obj in self.widgets_cu_cursor_mana:
            QApplication.restoreOverrideCursor()
            return True
        return super().eventFilter(obj, event)


# --- Bloc principal pentru rularea aplicației standalone ---

if __name__ == '__main__':
    if not os.path.exists(DB_MEMBRII) or not os.path.exists(DB_DEPCRED):  # MODIFICARE - Folosim căile definite
        msgBox = QMessageBox()
        msgBox.setIcon(QMessageBox.Critical)
        msgBox.setWindowTitle("Eroare Fișiere Lipsă")
        msgBox.setText("Baze de date MEMBRII.db sau DEPCRED.db lipsă!")
        msgBox.setInformativeText("Asigurați-vă că ambele fișiere .db se află în același director cu aplicația.")
        msgBox.setStandardButtons(QMessageBox.Ok)
        msgBox.exec_()
        sys.exit(1)

    app = QApplication(sys.argv)

    if "Fusion" in QtWidgets.QStyleFactory.keys():  # Folosim QtWidgets aici
        app.setStyle("Fusion")

    main_window = QWidget()
    main_window.setWindowTitle("Vizualizare Situație Lunară CAR")
    main_layout = QVBoxLayout(main_window)
    widget = VizualizareLunaraWidget()
    main_layout.addWidget(widget)
    main_window.setMinimumSize(950, 600)  # Dimensiune minimă similară
    main_window.show()

    sys.exit(app.exec_())